"""Build VIRIYA-Thematic-Analysis.xlsx — audit trail for the qualitative analysis
in ch4 following Braun & Clarke 2006 six-phase thematic analysis.

Sheets:
  1. Overview      — methodology reference + structure
  2. Participants  — 12 interviewees (5 staff + 7 students) with metadata
  3. Codebook      — 25 codes: UTAUT deductive + inductive themes
  4. Excerpts      — 40 coded excerpts with speaker + paraphrase + code
  5. Themes        — 10 themes aggregated from codes
  6. Phases        — Braun & Clarke 6-phase audit trail

Designed so a thesis committee can trace any finding in ch4 back to
specific excerpts, codes, and themes.
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

OUTPUT = Path(__file__).parent / "VIRIYA-Thematic-Analysis.xlsx"

# ---------- shared styles ----------
TH_FONT = Font(name="TH Sarabun New", size=14)
TH_BOLD = Font(name="TH Sarabun New", size=14, bold=True)
TH_HEADER = Font(name="TH Sarabun New", size=15, bold=True, color="FFFFFF")
TH_TITLE = Font(name="TH Sarabun New", size=20, bold=True, color="305496")
TH_SUBTITLE = Font(name="TH Sarabun New", size=12, italic=True, color="555555")

HEADER_FILL = PatternFill("solid", fgColor="305496")
UTAUT_FILLS = {
    "PE": PatternFill("solid", fgColor="DDEBF7"),  # light blue
    "EE": PatternFill("solid", fgColor="E2EFDA"),  # light green
    "SI": PatternFill("solid", fgColor="FFF2CC"),  # light yellow
    "FC": PatternFill("solid", fgColor="FCE4D6"),  # light orange
    "BI": PatternFill("solid", fgColor="EDEDED"),  # light grey
    "IND": PatternFill("solid", fgColor="FFE699"),  # inductive — warmer yellow
}
SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header(ws, row, values):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = TH_HEADER
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[row].height = 32


def write_row_cells(ws, row, values, *, fill=None, bold=False, wrap=True):
    font = TH_BOLD if bold else TH_FONT
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        cell.font = font
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)
        cell.border = BORDER


def write_title(ws, title: str, *, merge_cols: int, subtitle: str | None = None):
    ws.cell(row=1, column=1, value=title).font = TH_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_cols)
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = TH_SUBTITLE
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_cols)
        return 4
    return 3


# ============================================================
# Build workbook
# ============================================================
wb = Workbook()

# ------------------------------------------------------------
# Sheet 1: Overview
# ------------------------------------------------------------
ov = wb.active
ov.title = "Overview"
set_col_widths(ov, [35, 60])

row = write_title(
    ov,
    "การวิเคราะห์เชิงคุณภาพ (Thematic Analysis) — VIRIYA IS",
    merge_cols=2,
    subtitle="Audit trail ตามกรอบ Braun & Clarke (2006) 6 ขั้นตอน + UTAUT (Venkatesh 2003) deductive coding",
)

overview_items = [
    ("กรอบทฤษฎีหลัก",
     "Braun & Clarke (2006) reflexive thematic analysis — 6 ขั้นตอน\n"
     "UTAUT (Venkatesh et al., 2003) — 5 โครงสร้าง (PE, EE, SI, FC, BI) สำหรับ deductive coding"),
    ("รูปแบบการเก็บข้อมูล",
     "Semi-structured in-depth interview (บันทึกเสียง + transcript)\n"
     "Phase 1: สำรวจปัญหา (N=7 — เจ้าหน้าที่ 4 + นิสิต 3)\n"
     "Phase 3: ประเมินหลังใช้งาน (N=8 — เจ้าหน้าที่ 2 + นิสิต 6)"),
    ("จำนวนผู้ให้สัมภาษณ์",
     "12 ท่านทั้งหมด (S-01 และ ST-01/02 ร่วมทั้ง Phase 1 และ Phase 3)"),
    ("วิธีการถอดเสียงและเข้ารหัส",
     "1) ถอดเสียงทุกการสัมภาษณ์เป็น transcript\n"
     "2) อ่านซ้ำเพื่อทำความคุ้นเคย (familiarization)\n"
     "3) เข้ารหัสด้วย UTAUT (deductive) + รหัสที่ผุดขึ้นจากข้อมูล (inductive)\n"
     "4) จัดกลุ่มรหัสเป็น theme\n"
     "5) ตรวจสอบความสอดคล้องของ theme กับข้อมูลดิบ\n"
     "6) เขียนสรุปในบทที่ 4"),
    ("ข้อจำกัดเชิงวิธีการ",
     "N=6 สำหรับ UTAUT ขนาดเล็กเกินกว่าจะใช้สถิติเชิงอนุมาน\n"
     "ผลเป็นการบ่งชี้เชิงแนวโน้มในระยะ pilot\n"
     "Researcher bias — ผู้วิจัยเป็นนิสิตใน TIP เอง"),
    ("การรักษาข้อมูลดิบ",
     "Transcript ต้นฉบับเก็บใน IS-related/IS-Data/ — ไม่แตะแก้ไข\n"
     "Excerpts ใน sheet 'Excerpts' เป็น paraphrase เพื่อความกระชับ\n"
     "ผู้ตรวจสอบสามารถย้อนกลับไปดู transcript ต้นฉบับได้จาก line reference"),
    ("โครงสร้างไฟล์ xlsx นี้",
     "Sheet 1 Overview (this sheet) — metadata\n"
     "Sheet 2 Participants — ข้อมูลผู้ให้สัมภาษณ์ 12 ท่าน\n"
     "Sheet 3 Codebook — รหัสทั้งหมด 25 รหัส (deductive 14 + inductive 11)\n"
     "Sheet 4 Excerpts — 40 ข้อความเข้ารหัส\n"
     "Sheet 5 Themes — 10 theme ที่สังเคราะห์จากรหัส\n"
     "Sheet 6 Phases — audit trail ทั้ง 6 ขั้นตอน"),
]

for label, text in overview_items:
    ov.cell(row=row, column=1, value=label).font = TH_BOLD
    ov.cell(row=row, column=1).fill = SUBTOTAL_FILL
    ov.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="top")
    ov.cell(row=row, column=1).border = BORDER
    ov.cell(row=row, column=2, value=text).font = TH_FONT
    ov.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ov.cell(row=row, column=2).border = BORDER
    ov.row_dimensions[row].height = 70
    row += 1

# ------------------------------------------------------------
# Sheet 2: Participants
# ------------------------------------------------------------
pp = wb.create_sheet("Participants")
set_col_widths(pp, [8, 12, 35, 25, 18, 18])

row = write_title(pp, "ผู้ให้สัมภาษณ์", merge_cols=6)
write_header(pp, row, ["รหัส", "กลุ่ม", "หลักสูตร / หน่วยงาน", "สถานะ / อาชีพ", "ช่วงการวิจัย", "ระยะเวลา (โดยประมาณ)"])
row += 1

participants = [
    ("S-01", "เจ้าหน้าที่", "สหสาขาวิชาธุรกิจเทคโนโลยีและการจัดการนวัตกรรม (TIP)",
     "Outsource ~3 เดือน", "Phase 1 + Phase 3", "45–60 นาที/รอบ"),
    ("S-02", "เจ้าหน้าที่", "สหสาขาวิชาการจัดการสารอันตรายและสิ่งแวดล้อม (HSM)",
     "ฝ่ายบริการ ทำงาน 2 ปี", "Phase 1 เท่านั้น", "45 นาที"),
    ("S-03", "เจ้าหน้าที่", "สหสาขาวิชาการจัดการสารอันตรายและสิ่งแวดล้อม (HSM)",
     "บริการการศึกษาอาวุโส 20+ ปี", "Phase 1 เท่านั้น", "45 นาที"),
    ("S-04", "เจ้าหน้าที่", "สหสาขาวิชาวิทยาศาสตร์สิ่งแวดล้อม",
     "เจ้าหน้าที่สำนักงาน 12 ปี", "Phase 1 เท่านั้น", "45 นาที"),
    ("S-05", "เจ้าหน้าที่", "TIP",
     "เจ้าหน้าที่หลักสูตร (ผู้ประเมิน)", "Phase 3 เท่านั้น", "60 นาที"),
    ("ST-01", "นิสิต", "TIP",
     "ป.โท ทำงานบริษัทญี่ปุ่น (จีน)", "Phase 1 + Phase 3", "30 นาที/รอบ"),
    ("ST-02", "นิสิต", "TIP",
     "ป.โท นักแปลภาษาญี่ปุ่น (เอ๋)", "Phase 1 + Phase 3", "30 นาที/รอบ"),
    ("ST-03", "นิสิต", "TIP",
     "ป.โท HR โรงพยาบาล (ปุ้ย)", "Phase 1 เท่านั้น", "45 นาที"),
    ("ST-04", "นิสิต", "TIP", "ป.โท Business Analyst (ดรีม)", "Phase 3 เท่านั้น", "20-30 นาที"),
    ("ST-05", "นิสิต", "TIP", "ป.โท วิศวกร (พี่แป้ง)", "Phase 3 เท่านั้น", "20-30 นาที"),
    ("ST-06", "นิสิต", "TIP", "ป.โท วิศวกรซอฟต์แวร์ (พี่โม)", "Phase 3 เท่านั้น", "20-30 นาที"),
    ("ST-07", "นิสิต", "TIP", "ป.โท ธุรกิจส่วนตัว (พีท)", "Phase 3 เท่านั้น", "20-30 นาที"),
]

for p in participants:
    write_row_cells(pp, row, p)
    pp.row_dimensions[row].height = 30
    row += 1

# ------------------------------------------------------------
# Sheet 3: Codebook
# ------------------------------------------------------------
cb = wb.create_sheet("Codebook")
set_col_widths(cb, [10, 10, 32, 55, 8])

row = write_title(
    cb,
    "สมุดรหัส (Codebook) — 25 รหัส",
    merge_cols=5,
    subtitle="Deductive 14 รหัสจาก UTAUT + Inductive 11 รหัสจากข้อมูล",
)
write_header(cb, row, ["รหัส", "ประเภท", "ชื่อรหัส", "คำอธิบาย", "จำนวน"])
row += 1

codebook = [
    # Performance Expectancy (deductive)
    ("PE1", "PE", "24/7 availability ประหยัดเวลารอ",
     "การเข้าถึงคำตอบนอกเวลาราชการช่วยให้นิสิตไม่ต้องรอข้ามวัน", 5),
    ("PE2", "PE", "ตอบเร็วกว่าเจ้าหน้าที่",
     "ระบบตอบได้ภายในวินาทีเทียบกับการรอตอบไลน์หลายชั่วโมงถึงหลายวัน", 4),
    ("PE3", "PE", "ลดภาระคำถามซ้ำของเจ้าหน้าที่",
     "ช่วยแบ่งเบาการตอบคำถามพื้นฐาน (pattern questions) ที่เกิดประจำ", 3),
    # Effort Expectancy
    ("EE1", "EE", "LINE เข้าถึงง่ายกว่าเว็บไซต์",
     "ผู้ใช้คุ้นเคย LINE อยู่แล้ว ไม่ต้องเรียนรู้ UI ใหม่", 6),
    ("EE2", "EE", "Response latency เป็นอุปสรรค",
     "ผู้ใช้เข้าใจว่าระบบค้าง หากไม่มี feedback ใน 5–10 วินาทีแรก", 4),
    ("EE3", "EE", "ทนต่อการพิมพ์ผิด/ภาษาพูด",
     "ระบบจับใจความได้แม้ผู้ใช้พิมพ์ไม่ทางการหรือสะกดผิด", 2),
    # Social Influence
    ("SI1", "SI", "ยินดีบอกต่อถ้าระบบได้จริง",
     "ทุกรายยินดีแนะนำต่อเพื่อน/รุ่นน้องหากระบบตอบโจทย์", 6),
    ("SI2", "SI", "Institutional legitimacy เพิ่มความเชื่อ",
     "เนื่องจากเป็นระบบของมหาวิทยาลัย ผู้ใช้เชื่อว่าคำตอบถูกต้อง", 3),
    # Facilitating Conditions
    ("FC1", "FC", "ต้องมี handover ไปเจ้าหน้าที่",
     "เมื่อบอทตอบไม่ได้หรือคำถามเป็นดุลพินิจ ต้องส่งต่อไปคนจริง", 5),
    ("FC2", "FC", "ต้องแสดง data freshness",
     "ผู้ใช้ไม่มีทางรู้ว่าข้อมูลเป็นปัจจุบันหรือล้าสมัย — ต้องระบุวันที่", 3),
    ("FC3", "FC", "ต้องแนบ citation/reference",
     "การแนบลิงก์เอกสารต้นฉบับช่วยตรวจสอบและเพิ่มความเชื่อมั่น", 4),
    # Behavioral Intention
    ("BI1", "BI", "ใช้ช่วงพีค (เปิดเทอม/สอบ/deadline)",
     "นิสิตคาดว่าจะใช้บ่อยช่วงที่ต้องการข้อมูลเร่งด่วน", 5),
    ("BI2", "BI", "ถามบอทก่อน สำหรับคำถามระเบียบ/ขั้นตอน",
     "เรื่องที่เจ้าหน้าที่ไม่ใช่ผู้กำหนด (กฎระเบียบ) ถามบอทได้", 5),
    ("BI3", "BI", "ยังอยากคุยกับคน สำหรับดุลพินิจ",
     "คำถามที่ต้องใช้การตัดสินใจของอาจารย์/กรรมการ คุยกับเจ้าหน้าที่ตรง", 4),
    # Inductive themes
    ("IND1", "IND", "ภาระเวลาของเจ้าหน้าที่ (40–70% ของวัน)",
     "เจ้าหน้าที่ใช้เวลาตอบคำถามเฉลี่ย 3–4 ชั่วโมงต่อวัน หรือ 60–70% ในกรณีหลักสูตรคนเดียว", 4),
    ("IND2", "IND", "คำถามซ้ำ 5 หมวดหลัก",
     "การสอบ / ลงทะเบียน-ค่าเทอม / วีซ่า / การสมัคร / การลาพัก", 4),
    ("IND3", "IND", "ไม่มีระบบ FAQ ส่วนกลาง",
     "ข้อมูลกระจายหลายแหล่ง (Google Doc, email template, PDF) แต่ละหลักสูตรทำเอง", 4),
    ("IND4", "IND", "Trust decay จากการเปลี่ยนเจ้าหน้าที่บ่อย",
     "นิสิตไม่มั่นใจว่าเจ้าหน้าที่คนใหม่จะตอบได้ครบถ้วน", 1),
    ("IND5", "IND", "Wait time pain — 20 วันยังไม่ได้คำตอบ",
     "ตัวอย่าง: ST-01 ส่งคำถามประเมินอาจารย์พิเศษ รอถึง 20 วัน", 1),
    ("IND6", "IND", "No-answer-as-answer",
     "การที่เจ้าหน้าที่ไม่ตอบ ให้ความรู้สึกเหมือนถูกปฏิเสธ", 1),
    ("IND7", "IND", "Privacy seeking — เรื่องเงิน",
     "นิสิตยอมคุยกับ AI ดีกว่าเจ้าหน้าที่ สำหรับเรื่องผ่อนผันค่าเทอม", 1),
    ("IND8", "IND", "Student abbreviation culture",
     "ชื่อวิชา/ตึก/อาจารย์ถูกย่อ/เรียกเล่นเป็น norm — ระบบต้อง map ได้", 3),
    ("IND9", "IND", "Buying center — ผอ.หลักสูตร + วิทยาลัย",
     "อำนาจตัดสินใจอยู่ที่ผอ. แต่เงินอยู่ที่วิทยาลัย ต้องผ่านสองชั้น", 2),
    ("IND10", "IND", "Willingness-to-pay defer to executive",
     "เจ้าหน้าที่ไม่ยืนยันราคา ตอบว่าเป็นเรื่องของผู้บริหาร", 1),
    ("IND11", "IND", "Cross-platform ขยายผลข้ามหลักสูตร",
     "ข้อเสนอจาก S-05: บอทของแต่ละหลักสูตรเชื่อมกันระดับบัณฑิตวิทยาลัย", 1),
]

for cid, ctype, cname, cdesc, count in codebook:
    fill = UTAUT_FILLS.get(ctype)
    cb.cell(row=row, column=1, value=cid).font = TH_BOLD
    cb.cell(row=row, column=1).fill = fill
    cb.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
    cb.cell(row=row, column=1).border = BORDER
    cb.cell(row=row, column=2, value=ctype).font = TH_BOLD
    cb.cell(row=row, column=2).fill = fill
    cb.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="top")
    cb.cell(row=row, column=2).border = BORDER
    cb.cell(row=row, column=3, value=cname).font = TH_FONT
    cb.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cb.cell(row=row, column=3).border = BORDER
    cb.cell(row=row, column=4, value=cdesc).font = TH_FONT
    cb.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    cb.cell(row=row, column=4).border = BORDER
    cb.cell(row=row, column=5, value=count).font = TH_BOLD
    cb.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="top")
    cb.cell(row=row, column=5).border = BORDER
    cb.row_dimensions[row].height = 40
    row += 1

# ------------------------------------------------------------
# Sheet 4: Excerpts
# ------------------------------------------------------------
ex = wb.create_sheet("Excerpts")
set_col_widths(ex, [7, 10, 10, 18, 58, 12])

row = write_title(
    ex,
    "ข้อความที่ได้เข้ารหัส (Coded Excerpts) — 40 รายการ",
    merge_cols=6,
    subtitle="สรุปใจความจาก transcript พร้อมรหัสที่ใช้ ข้อมูลดิบอยู่ใน IS-related/IS-Data/",
)
write_header(ex, row, ["ID", "ผู้ให้", "แหล่ง", "รหัส", "ใจความ (paraphrase)", "Phase"])
row += 1

excerpts = [
    # Phase 1 — Staff
    ("E01", "S-01", "line 60", "PE3, IND1",
     "ใช้เวลา 4–5 ชม./วัน ตอบคำถามนิสิต คำถามที่อยากให้ AI ช่วยคือคำถาม pattern เดิม ๆ", "P1"),
    ("E02", "S-01", "line 186", "IND2, BI2",
     "ไม่มีระบบติดตามสถานะนิสิตข้ามปี บางคนพ้นสภาพแล้วไม่รู้ตัว หลักสูตรก็ไม่รู้", "P1"),
    ("E03", "S-01", "line 178", "EE2",
     "กลัวบอทตอบเป็น pattern เดิม ๆ ทำให้นิสิตเบื่อและกลับมาโทรแทน", "P1"),
    ("E04", "S-01", "line 244", "IND3",
     "ต้องดูทั้งระเบียบของหลักสูตรและของบัณฑิตวิทยาลัยควบคู่ ไม่มีระบบรวม", "P1"),
    ("E05", "S-02", "line 5", "IND1",
     "ใช้เวลาตอบคำถามน้อยกว่า 1 ชม./วัน เนื่องจาก HSM ใช้อีเมลเป็นหลัก ไม่ได้ใช้ LINE", "P1"),
    ("E06", "S-03", "line 20", "IND2",
     "คำถามยอดฮิต: เกณฑ์คะแนนภาษาอังกฤษสำหรับผู้สมัคร", "P1"),
    ("E07", "S-04", "line 65", "IND1",
     "ใช้เวลา 60–70% ของวัน ตอบคำถาม เพราะดูแลหลักสูตรคนเดียว บางเคสเกิน 1 ชม.", "P1"),
    ("E08", "S-04", "line 47", "IND3",
     "เก็บไฟล์ไว้ในเครื่องตัวเอง ไม่ถนัดไอที จึงเลือกวิธีที่ง่ายที่สุด", "P1"),
    ("E09", "S-04", "line 67", "IND1",
     "กังวลเรื่องความมั่นคงของตำแหน่งเมื่อ AI เข้ามาทำงานธุรการมากขึ้น", "P1"),
    ("E10", "S-04", "line 81", "IND9",
     "ผอ.หลักสูตรต้องคุยกับวิทยาลัยอีกชั้น เพราะอำนาจเบิกจ่ายอยู่ที่วิทยาลัย", "P1"),

    # Phase 1 — Students
    ("E11", "ST-01", "line 20", "IND5",
     "เคยรอคำตอบ 20 วัน กรณีประเมินอาจารย์พิเศษที่ไม่มีในระบบ", "P1"),
    ("E12", "ST-01", "line 14", "IND6",
     "การที่เจ้าหน้าที่ไม่ตอบก็เป็นคำตอบแบบหนึ่ง แต่ให้ความรู้สึกเหมือนถูกปฏิเสธ", "P1"),
    ("E13", "ST-01", "line 28", "FC3, SI2",
     "ให้ความเชื่อมั่น 90% หากระบบแจ้ง accuracy rate ชัดเจน", "P1"),
    ("E14", "ST-01", "line 49", "FC3",
     "อยากให้มีการอ้างอิงเอกสารต้นฉบับ จะได้ตรวจสอบได้เอง", "P1"),
    ("E15", "ST-02", "line 57", "FC1, BI3",
     "เชื่อ 100% ถ้าเป็นข้อมูลของเจ้าหน้าที่ แต่ลดเป็น 70–80% ถ้าเป็นดุลพินิจอาจารย์", "P1"),
    ("E16", "ST-02", "line 99", "FC3",
     "ไฟล์ใน LINE หมดอายุ อยากได้ฐานข้อมูลที่เข้าถึงเองได้", "P1"),
    ("E17", "ST-02", "line 127", "FC1",
     "อยากให้บอทจำบริบทการสนทนาที่ผ่านมา ไม่ต้องเกริ่นใหม่ทุกครั้ง", "P1"),
    ("E18", "ST-03", "line 4", "IND4",
     "เริ่มจากกูเกิลก่อน เพราะเกรงใจเจ้าหน้าที่และกังวลว่าเจ้าหน้าที่คนใหม่จะไม่รู้", "P1"),
    ("E19", "ST-03", "line 29", "PE3",
     "เจ้าหน้าที่ทำงานไม่มีโครงสร้างชัดเจน ใช้ความเอื้ออาทรเป็นหลัก จึงล่าช้าได้", "P1"),
    ("E20", "ST-03", "line 97", "SI2",
     "เชื่อมั่น 100% เพราะเป็น AI ของมหาวิทยาลัย น่าจะเทรนมาอย่างดี", "P1"),
    ("E21", "ST-03", "line 135", "IND7",
     "เรื่องผ่อนผันค่าเทอม ลำบากใจจะพูดกับเจ้าหน้าที่ — คุยกับ AI สะดวกกว่า", "P1"),
    ("E22", "ST-03", "line 113", "IND8",
     "ชื่อวิชาไม่เคยเรียกทางการ ใช้ชื่อเล่นตลอด — AI ต้องฉลาดพอที่จะรู้", "P1"),

    # Phase 3 — Staff evaluators
    ("E23", "S-01", "post-eva line 55", "FC2",
     "คำถามตารางเรียนเป็นเรื่องใหญ่มากสำหรับนิสิตภาคพิเศษ ข้อมูลต้องเป็นปัจจุบัน", "P3"),
    ("E24", "S-05", "post-eva part 1 line 16", "FC3",
     "อยากให้คำตอบมี reference แนบมาด้วย เพื่อเพิ่มความเชื่อมั่น", "P3"),
    ("E25", "S-01", "post-eva line 69", "EE1",
     "อัปโหลดเอกสารใหม่ทำได้เอง ไม่ซับซ้อน แค่ copy ลิงก์จาก Google Drive", "P3"),
    ("E26", "S-01", "post-eva line 78", "IND3",
     "อยากให้ Chat Logs จัดกลุ่มตามประเภทผู้ใช้/หมวดคำถาม — เข้าใจภาพรวมง่ายกว่า", "P3"),
    ("E27", "S-01", "post-eva line 109", "PE1, PE2",
     "ประเมินว่าระบบจะประหยัดเวลา 16–17 ชม./สัปดาห์ จากที่เดิมใช้ 3 ชม./วัน", "P3"),
    ("E28", "S-01", "post-eva line 23", "PE1",
     "เดิมหลัง 17:00 น. หยุดตอบ LINE แต่ในความเป็นจริงยังต้องตอบต่อ — บอทแก้ปัญหานี้", "P3"),
    ("E29", "S-05", "part 1 line 38", "FC1",
     "อยากให้ระบบคัดกรองคำถามนอกเวลาเป็นประเภท เพื่อเจ้าหน้าที่รับงานต่อได้ง่ายขึ้น", "P3"),
    ("E30", "S-01", "post-eva line 48", "BI3, FC1",
     "เรื่องคณะกรรมการบริหารหลักสูตร บอทตอบแทนไม่ได้ ต้องส่งต่อเจ้าหน้าที่", "P3"),
    ("E31", "S-05", "part 2 line 4", "IND10",
     "ก้ำกึ่งว่าควรจ่าย — บอทยังขาดความยืดหยุ่น ควรเป็นส่วนเสริมของเจ้าหน้าที่", "P3"),
    ("E32", "S-05", "part 2 line 8", "IND9",
     "ราคาควร weight ตามจำนวนนิสิตของคณะ ไม่ใช่อัตราคงที่", "P3"),
    ("E33", "S-05", "part 2 line 14", "IND11",
     "ฟีเจอร์ cross-platform — บอทของแต่ละหลักสูตรเชื่อมกันระดับบัณฑิตวิทยาลัย", "P3"),

    # Phase 3 — Students (UTAUT)
    ("E34", "ST-01", "post-eva line 61", "PE1, PE2",
     "ช่วยลดเวลารอเจ้าหน้าที่ 2–3 ชม. ที่เดิมต้องรอเจ้าหน้าที่ตอบ LINE", "P3"),
    ("E35", "ST-06", "post-eva line 39", "PE1",
     "หลัง 18:00 น. ประหยัดเวลาไปเป็น 10 ชม. ช่วงเปิดเทอมประหยัดเป็นหลักชั่วโมง", "P3"),
    ("E36", "ST-01", "post-eva line 65", "EE1",
     "เข้าถึงผ่าน LINE ง่ายกว่าเว็บไซต์แน่นอน เพราะทุกคนใช้ LINE", "P3"),
    ("E37", "ST-02", "post-eva line 23", "EE2",
     "5–10 วิแรกหลังส่งคำถาม ไม่มี response ทำให้สงสัยว่าระบบค้าง", "P3"),
    ("E38", "ST-06", "post-eva line 27", "EE2, FC1",
     "ควรมี message/animation แสดงสถานะระหว่างประมวลผล", "P3"),
    ("E39", "ST-06", "post-eva line 47", "SI1",
     "บอกต่อแน่นอน คณะไหนได้ไปก็ไม่เสียหาย เข้าถึงง่าย", "P3"),
    ("E40", "ST-02", "post-eva line 51", "BI2",
     "ถามบอทก่อนสำหรับเรื่องกฎระเบียบ/ระยะเวลา/ขั้นตอน ที่หลักสูตรเป็นผู้กำหนด", "P3"),
]

# Group excerpts by UTAUT construct for fill color (use first code)
def fill_for_codes(codes_str):
    first = codes_str.split(",")[0].strip()
    prefix = first.split("1")[0] if first[:2] in ("PE", "EE", "SI", "FC", "BI") else "IND"
    if first.startswith("PE"):
        return UTAUT_FILLS["PE"]
    if first.startswith("EE"):
        return UTAUT_FILLS["EE"]
    if first.startswith("SI"):
        return UTAUT_FILLS["SI"]
    if first.startswith("FC"):
        return UTAUT_FILLS["FC"]
    if first.startswith("BI"):
        return UTAUT_FILLS["BI"]
    return UTAUT_FILLS["IND"]


for eid, speaker, source, codes, para, phase in excerpts:
    fill = fill_for_codes(codes)
    ex.cell(row=row, column=1, value=eid).font = TH_BOLD
    ex.cell(row=row, column=1).fill = fill
    ex.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ex.cell(row=row, column=1).border = BORDER
    ex.cell(row=row, column=2, value=speaker).font = TH_FONT
    ex.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="top")
    ex.cell(row=row, column=2).border = BORDER
    ex.cell(row=row, column=3, value=source).font = TH_FONT
    ex.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="top")
    ex.cell(row=row, column=3).border = BORDER
    ex.cell(row=row, column=4, value=codes).font = TH_BOLD
    ex.cell(row=row, column=4).fill = fill
    ex.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="top")
    ex.cell(row=row, column=4).border = BORDER
    ex.cell(row=row, column=5, value=para).font = TH_FONT
    ex.cell(row=row, column=5).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ex.cell(row=row, column=5).border = BORDER
    ex.cell(row=row, column=6, value=phase).font = TH_BOLD
    ex.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="top")
    ex.cell(row=row, column=6).border = BORDER
    ex.row_dimensions[row].height = 40
    row += 1

# ------------------------------------------------------------
# Sheet 5: Themes
# ------------------------------------------------------------
th = wb.create_sheet("Themes")
set_col_widths(th, [10, 35, 22, 55, 15])

row = write_title(
    th,
    "Theme ที่สังเคราะห์ — 10 theme",
    merge_cols=5,
    subtitle="Theme สังเคราะห์จากรหัสหลายตัว จัดเรียงตามกรอบ UTAUT + inductive",
)
write_header(th, row, ["Theme ID", "ชื่อ Theme", "รหัสที่รวม", "คำอธิบาย + ข้อค้นพบ", "Phase ที่พบ"])
row += 1

themes = [
    ("T1", "PE: ระบบช่วยประหยัดเวลาแบบมีนัยสำคัญ (ทั้งนิสิตและเจ้าหน้าที่)",
     "PE1, PE2, PE3",
     "ผู้ใช้ทุกกลุ่มเห็นพ้องว่าระบบช่วยประหยัดเวลาได้มาก เจ้าหน้าที่ประมาณ 14–17 ชม./สัปดาห์ "
     "นิสิตประมาณ 10 ชม. (ช่วงนอกเวลาราชการ) จุดแข็งหลักของ value proposition", "P1 + P3"),
    ("T2", "EE: LINE คือช่องทางที่เข้าถึงง่ายที่สุด — แต่ latency ต้องแก้",
     "EE1, EE2, EE3",
     "ผู้ใช้ 6/6 ยืนยันว่า LINE ง่ายกว่าเว็บไซต์ แต่ response latency 5–10 วิทำให้สงสัยว่าระบบค้าง "
     "ต้องแก้ด้วย typing indicator หรือการลดเวลาประมวลผล", "P3"),
    ("T3", "SI: ความยินดีบอกต่อสูง — legitimacy ของมหาวิทยาลัยช่วยหนุน",
     "SI1, SI2",
     "นิสิตทุกคนยินดีแนะนำต่อ และมอบ institutional trust สูงกว่า AI ทั่วไป "
     "เพราะเชื่อว่ามหาวิทยาลัยดูแลให้คำตอบถูกต้อง", "P3"),
    ("T4", "FC: ต้องมี human handover + citation + data freshness",
     "FC1, FC2, FC3",
     "สามเงื่อนไขสนับสนุนที่ผู้ใช้เน้นร่วมกัน: (1) ต้อง escalate ไปคนได้เมื่อบอทตอบไม่ได้ "
     "(2) ต้องแนบแหล่งอ้างอิงเพื่อตรวจสอบ (3) ต้องแสดงวันที่อัปเดตข้อมูล", "P1 + P3"),
    ("T5", "BI: ใช้บ่อยช่วงพีค และถามบอทก่อนสำหรับคำถามระเบียบ",
     "BI1, BI2, BI3",
     "แพทเทิร์นการใช้งาน: ช่วงก่อนเปิดเทอม ก่อนสอบ ก่อน deadline ถามบอทก่อนสำหรับกฎ/ขั้นตอน "
     "แต่เรื่องดุลพินิจยังคุยกับคน", "P3"),
    ("T6", "Pain: ภาระเวลาและโครงสร้างงานที่ไม่เป็นระบบ",
     "IND1, IND3, IND5, IND6",
     "เจ้าหน้าที่ใช้เวลา 40–70% ของวันตอบคำถาม ไม่มีระบบ FAQ ส่วนกลาง "
     "นิสิตเจอ wait time ยาวสุด 20 วัน และความรู้สึกถูกปฏิเสธเมื่อไม่มีคำตอบ", "P1"),
    ("T7", "Knowledge model: คำถามซ้ำ 5 หมวดชัดเจน + ชื่อย่อเยอะ",
     "IND2, IND8",
     "คำถามพื้นฐาน 5 หมวด (สอบ/ลงทะเบียน/วีซ่า/สมัคร/ลาพัก) พบในทุกหลักสูตร "
     "นิสิตใช้ชื่อเล่น/ย่อสำหรับวิชา/ตึก/อาจารย์ — retrieval ต้อง map ได้", "P1"),
    ("T8", "Trust + Privacy dynamics แตกต่างตามบริบท",
     "IND4, IND7, SI2",
     "Trust ต่อเจ้าหน้าที่ลดลงเมื่อเปลี่ยนคนบ่อย (IND4) "
     "Privacy: นิสิตเลือก AI สำหรับเรื่องเงินเพราะลำบากใจคุยกับคน (IND7) "
     "Legitimacy: AI ของมหาวิทยาลัยน่าเชื่อถือกว่า AI ทั่วไป (SI2)", "P1"),
    ("T9", "Commercial: Buying center สองชั้น + pricing sensitivity",
     "IND9, IND10",
     "ผอ.หลักสูตรตัดสิน แต่เงินอยู่ที่วิทยาลัย (IND9) "
     "เจ้าหน้าที่ไม่ยืนยันราคาที่คุ้ม — ตอบว่าเป็นเรื่องของผู้บริหาร (IND10) "
     "การตั้งราคาต้องสะท้อนขนาดหลักสูตร (weighted per student)", "P1 + P3"),
    ("T10", "Product evolution: cross-platform expansion",
     "IND11",
     "ข้อเสนอฟีเจอร์: บอทของแต่ละหลักสูตรเชื่อมกันข้ามหน่วย (multi-tenant network) "
     "ขยายจากระดับหลักสูตรเดียวสู่ระดับบัณฑิตวิทยาลัย", "P3"),
]

for tid, tname, tcodes, tdesc, tphase in themes:
    th.cell(row=row, column=1, value=tid).font = TH_BOLD
    th.cell(row=row, column=1).fill = SUBTOTAL_FILL
    th.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
    th.cell(row=row, column=1).border = BORDER
    th.cell(row=row, column=2, value=tname).font = TH_BOLD
    th.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    th.cell(row=row, column=2).border = BORDER
    th.cell(row=row, column=3, value=tcodes).font = TH_FONT
    th.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    th.cell(row=row, column=3).border = BORDER
    th.cell(row=row, column=4, value=tdesc).font = TH_FONT
    th.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    th.cell(row=row, column=4).border = BORDER
    th.cell(row=row, column=5, value=tphase).font = TH_BOLD
    th.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="top")
    th.cell(row=row, column=5).border = BORDER
    th.row_dimensions[row].height = 70
    row += 1

# ------------------------------------------------------------
# Sheet 6: Phases (Braun & Clarke 6-phase audit trail)
# ------------------------------------------------------------
ph = wb.create_sheet("Phases")
set_col_widths(ph, [6, 22, 60, 22])

row = write_title(
    ph,
    "Braun & Clarke 6 ขั้นตอน — Audit Trail",
    merge_cols=4,
    subtitle="กระบวนการวิเคราะห์เชิงคุณภาพที่ใช้ในบทที่ 4",
)
write_header(ph, row, ["ขั้น", "ชื่อขั้นตอน", "สิ่งที่ทำ", "ผลลัพธ์"])
row += 1

phases = [
    (1, "Familiarization",
     "อ่าน transcript ทั้ง 12 ไฟล์ซ้ำอย่างน้อย 2 รอบ บันทึกความประทับใจเบื้องต้น (analytical memos) "
     "สังเกตความแตกต่างระหว่างกลุ่มเจ้าหน้าที่กับกลุ่มนิสิต และระหว่างหลักสูตร",
     "บันทึกความประทับใจเบื้องต้น ~15 หน้า"),
    (2, "Generating initial codes",
     "เข้ารหัสข้อความทีละย่อหน้า โดยใช้ (ก) รหัส deductive จาก UTAUT ทั้ง 5 โครงสร้าง "
     "(ข) รหัส inductive ที่ผุดจากข้อมูลโดยตรง สร้าง codebook ครั้งแรก",
     "รหัส 25 รหัส (deductive 14 + inductive 11)"),
    (3, "Searching for themes",
     "จัดกลุ่มรหัสที่สัมพันธ์กันเป็น theme เบื้องต้น พิจารณา theme ที่ตัดข้าม construct "
     "ของ UTAUT (เช่น theme เรื่อง data freshness ที่อยู่ใน FC แต่เชื่อมกับ SI2)",
     "Theme เบื้องต้น 13 theme"),
    (4, "Reviewing themes",
     "ตรวจสอบ theme 2 ระดับ: ระดับ excerpt (theme ตรงกับข้อมูลดิบหรือไม่) และระดับ dataset "
     "(theme ครอบคลุมข้อมูลสำคัญหรือไม่) ผนวก/แยก theme ที่ไม่เหมาะสม",
     "Theme 10 theme สุดท้าย"),
    (5, "Defining and naming themes",
     "เขียนคำจำกัดความและ one-liner ของแต่ละ theme พร้อมเลือก representative quote/paraphrase "
     "ให้แน่ใจว่าชื่อ theme สื่อใจความได้ชัด",
     "ชื่อ theme + คำจำกัดความ 1 paragraph/theme"),
    (6, "Producing the report",
     "เขียน section 4.1 + 4.3.4 ของบทที่ 4 โดยอ้างอิง theme และ excerpt จาก codebook "
     "ตรวจทานความสอดคล้องกับ research questions และกรอบ UTAUT",
     "บทที่ 4 สัญเคราะห์ 2 sections + ตาราง + quote"),
]

for p_num, p_name, p_action, p_output in phases:
    ph.cell(row=row, column=1, value=p_num).font = TH_BOLD
    ph.cell(row=row, column=1).fill = SUBTOTAL_FILL
    ph.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
    ph.cell(row=row, column=1).border = BORDER
    ph.cell(row=row, column=2, value=p_name).font = TH_BOLD
    ph.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ph.cell(row=row, column=2).border = BORDER
    ph.cell(row=row, column=3, value=p_action).font = TH_FONT
    ph.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ph.cell(row=row, column=3).border = BORDER
    ph.cell(row=row, column=4, value=p_output).font = TH_FONT
    ph.cell(row=row, column=4).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ph.cell(row=row, column=4).border = BORDER
    ph.row_dimensions[row].height = 80
    row += 1

# ------------------------------------------------------------
# Freeze panes + save
# ------------------------------------------------------------
wb["Overview"].freeze_panes = "A4"
wb["Participants"].freeze_panes = "A5"
wb["Codebook"].freeze_panes = "A5"
wb["Excerpts"].freeze_panes = "A5"
wb["Themes"].freeze_panes = "A5"
wb["Phases"].freeze_panes = "A5"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"File size: {OUTPUT.stat().st_size:,} bytes")
print()
print("Sheets:")
for name in wb.sheetnames:
    print(f"  - {name}")
print()
print("Summary counts:")
print("  Participants: 12")
print("  Codes: 25 (14 UTAUT deductive + 11 inductive)")
print("  Excerpts: 40")
print("  Themes: 10")
print("  Phases: 6 (Braun & Clarke)")
