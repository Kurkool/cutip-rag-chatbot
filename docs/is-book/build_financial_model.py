"""Build VIRIYA-Financial-Model.xlsx — 5-year projection with live Excel formulas.

Sheets (follow ch6 flow):
  1. Assumptions    — §6.2 all inputs, yellow = editable, incl. loan + dividend
  2. Loan           — §6.6.3 amortization schedule 5 years
  3. Assets         — §6.1 asset list + depreciation
  4. SG&A           — §6.5 selling + admin expense 5-year
  5. P&L            — §6.8 profit-and-loss 5-year incl. interest + dividend
  6. CashFlow       — §6.9 cash flow 5-year incl. financing
  7. BalanceSheet   — §6.7 year-end balance 5-year incl. loan
  8. BreakEven      — §6.10
  9. Sensitivity    — §6.11
  10. Summary       — §6.12 NPV / IRR / MIRR / Payback

Base-case numbers match ch06-financial-feasibility.md tables.
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

OUTPUT = Path(__file__).parent / "VIRIYA-Financial-Model.xlsx"

# ---------- styles ----------
TH_FONT = Font(name="TH Sarabun New", size=14)
TH_BOLD = Font(name="TH Sarabun New", size=14, bold=True)
TH_HEADER = Font(name="TH Sarabun New", size=16, bold=True, color="FFFFFF")
TH_SUBTITLE = Font(name="TH Sarabun New", size=13, italic=True, color="555555")

HEADER_FILL = PatternFill("solid", fgColor="305496")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
TOTAL_FILL = PatternFill("solid", fgColor="8EA9DB")

THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FMT_THB = '#,##0;[Red](#,##0)'
FMT_PCT = '0.0%'
FMT_INT = '#,##0'


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_cell(cell, font=TH_FONT, fill=None, fmt=None, align="right"):
    cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = BORDER


def write_row(ws, row, values, *, font=TH_FONT, fill=None, fmt=FMT_THB,
              label_align="left", value_align="right"):
    for col, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=v)
        align = label_align if col == 1 else value_align
        use_fmt = None if col == 1 else fmt
        style_cell(cell, font=font, fill=fill, fmt=use_fmt, align=align)


def write_header(ws, row, values):
    write_row(ws, row, values, font=TH_HEADER, fill=HEADER_FILL,
              fmt=None, label_align="center", value_align="center")
    ws.row_dimensions[row].height = 28


def write_label_row(ws, row, label, formulas, *, font=TH_FONT, fill=None, fmt=FMT_THB):
    ws.cell(row=row, column=1, value=label).font = font
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=1).border = BORDER
    if fill:
        ws.cell(row=row, column=1).fill = fill
    for i, f in enumerate(formulas, 2):
        cell = ws.cell(row=row, column=i, value=f)
        style_cell(cell, font=font, fill=fill, fmt=fmt, align="right")


def section_banner(ws, row, label, n_cols=6):
    ws.cell(row=row, column=1, value=label).font = TH_BOLD
    ws.cell(row=row, column=1).fill = SUBTOTAL_FILL
    ws.cell(row=row, column=1).border = BORDER
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    for i in range(2, n_cols + 1):
        ws.cell(row=row, column=i).fill = SUBTOTAL_FILL
        ws.cell(row=row, column=i).border = BORDER


wb = Workbook()

# ============================================================
# Sheet 1: Assumptions
# ============================================================
ws = wb.active
ws.title = "Assumptions"
set_col_widths(ws, [45, 18, 40])

ws.cell(row=1, column=1, value="VIRIYA Financial Model — Assumptions (5-year)").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
ws.cell(row=2, column=1, value="แก้ช่องเหลือง → P&L / Cash / BS / Summary recalculate อัตโนมัติ").font = TH_SUBTITLE
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

row = 4
A = {}

# --- Revenue assumptions ---
write_header(ws, row, ["รายการ (รายได้)", "ค่า", "หมายเหตุ"])
row += 1
revenue_rows = [
    ("price_starter",    "Starter tier — ราคา/เดือน (THB)",  6500,  "≤100 นิสิต"),
    ("price_pro",        "Pro tier — ราคา/เดือน (THB)",      15000, "100-300 นิสิต"),
    ("price_enterprise", "Enterprise tier — ราคา/เดือน (THB)", 32500, ">300 นิสิต"),
    ("mix_starter",      "สัดส่วน Starter",                   0.60,  ""),
    ("mix_pro",          "สัดส่วน Pro",                       0.30,  ""),
    ("mix_enterprise",   "สัดส่วน Enterprise",                0.10,  ""),
    ("setup_fee",        "Setup fee ต่อ tenant (THB)",        25000, "ครั้งเดียว"),
]
for key, label, value, note in revenue_rows:
    ws.cell(row=row, column=1, value=label).font = TH_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=1).border = BORDER
    fmt = FMT_PCT if "mix_" in key else FMT_THB
    style_cell(ws.cell(row=row, column=2, value=value),
               font=TH_BOLD, fill=INPUT_FILL, fmt=fmt)
    ws.cell(row=row, column=3, value=note).font = TH_SUBTITLE
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=3).border = BORDER
    A[key] = f"Assumptions!$B${row}"
    row += 1

# Weighted avg
ws.cell(row=row, column=1, value="ค่าเฉลี่ยถ่วงน้ำหนัก/tenant/mo").font = TH_BOLD
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
ws.cell(row=row, column=1).border = BORDER
style_cell(ws.cell(row=row, column=2,
                   value=f"={A['price_starter']}*{A['mix_starter']}"
                         f"+{A['price_pro']}*{A['mix_pro']}"
                         f"+{A['price_enterprise']}*{A['mix_enterprise']}"),
           font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
ws.cell(row=row, column=3, value="auto").font = TH_SUBTITLE
ws.cell(row=row, column=3).border = BORDER
A["avg_revenue"] = f"Assumptions!$B${row}"
row += 2

# --- Cost assumptions ---
write_header(ws, row, ["รายการ (ต้นทุน)", "ค่า", "หมายเหตุ"])
row += 1
cost_rows = [
    ("var_opus_in",    "Claude Opus input — ฿/tenant/mo",  1050,  "$5/Mtok × 6M"),
    ("var_opus_out",   "Claude Opus output — ฿/tenant/mo", 1313,  "$25/Mtok × 1.5M"),
    ("var_cohere_rr",  "Cohere Rerank — ฿/tenant/mo",      210,   ""),
    ("var_cohere_emb", "Cohere Embed — ฿/tenant/mo",       0.42,  ""),
    ("var_pinecone",   "Pinecone read — ฿/tenant/mo",      8,     ""),
    ("fix_pinecone",   "Pinecone Standard plan",           1750,  "$50/mo"),
    ("fix_gcp_run",    "GCP Cloud Run",                    1050,  "$30/mo"),
    ("fix_gcp_other",  "GCP Firestore + CS",               350,   ""),
    ("fix_other",      "Domain + SSL + monitoring",        350,   ""),
    ("salary_tech",    "Tech lead / คน/เดือน",              80000, ""),
    ("salary_biz",     "Customer/biz / คน/เดือน",           40000, ""),
    ("onboard_cost",   "ต้นทุน onboard ใหม่ (ครั้งเดียว)",   5000,  ""),
]
for key, label, value, note in cost_rows:
    ws.cell(row=row, column=1, value=label).font = TH_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=1).border = BORDER
    style_cell(ws.cell(row=row, column=2, value=value),
               font=TH_BOLD, fill=INPUT_FILL, fmt=FMT_THB)
    ws.cell(row=row, column=3, value=note).font = TH_SUBTITLE
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=3).border = BORDER
    A[key] = f"Assumptions!$B${row}"
    row += 1

# Computed totals
ws.cell(row=row, column=1, value="รวมต้นทุนผันแปร/tenant/mo").font = TH_BOLD
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
ws.cell(row=row, column=1).border = BORDER
style_cell(ws.cell(row=row, column=2,
                   value=f"={A['var_opus_in']}+{A['var_opus_out']}+{A['var_cohere_rr']}+{A['var_cohere_emb']}+{A['var_pinecone']}"),
           font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
ws.cell(row=row, column=3).border = BORDER
A["var_total"] = f"Assumptions!$B${row}"
row += 1

ws.cell(row=row, column=1, value="รวมต้นทุนคงที่ระบบ/mo").font = TH_BOLD
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
ws.cell(row=row, column=1).border = BORDER
style_cell(ws.cell(row=row, column=2,
                   value=f"={A['fix_pinecone']}+{A['fix_gcp_run']}+{A['fix_gcp_other']}+{A['fix_other']}"),
           font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
ws.cell(row=row, column=3).border = BORDER
A["fix_total"] = f"Assumptions!$B${row}"
row += 1

ws.cell(row=row, column=1, value="รวมเงินเดือนทีมเริ่มต้น/mo").font = TH_BOLD
ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
ws.cell(row=row, column=1).border = BORDER
style_cell(ws.cell(row=row, column=2, value=f"={A['salary_tech']}+{A['salary_biz']}"),
           font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
ws.cell(row=row, column=3).border = BORDER
A["salary_total"] = f"Assumptions!$B${row}"
row += 2

# --- Growth, finance, dividend ---
write_header(ws, row, ["รายการ (การเติบโต/ภาษี/เงินกู้/ปันผล)", "ค่า", "หมายเหตุ"])
row += 1
growth_rows = [
    ("growth_y1",  "Tenant ใหม่/mo ปีที่ 1",  2,       ""),
    ("growth_y2",  "Tenant ใหม่/mo ปีที่ 2",  2,       ""),
    ("growth_y3",  "Tenant ใหม่/mo ปีที่ 3",  1,       ""),
    ("growth_y4",  "Tenant ใหม่/mo ปีที่ 4",  1,       ""),
    ("growth_y5",  "Tenant ใหม่/mo ปีที่ 5",  1,       ""),
    ("tax_rate",   "อัตราภาษีนิติบุคคล",       0.20,    "CIT 20%"),
    ("wacc",       "WACC (คิดลด NPV)",         0.12,    "equity 67% × 15% + debt 33% × 4%"),
    ("founder_cap",    "เงินทุนผู้ถือหุ้น (THB)",       2000000, "2,000 หุ้น × 1,000 × 4 คน"),
    ("loan_amount",    "เงินกู้ระยะยาว (THB)",           1000000, "ภาครัฐ"),
    ("loan_rate",      "อัตราดอกเบี้ยเงินกู้",            0.05,    "5% ต่อปี"),
    ("loan_term",      "ระยะเวลาผ่อน (ปี)",              5,       ""),
    ("dividend_pct",   "นโยบายปันผล (ของกำไรสุทธิ)",      0.30,    "30%"),
    ("dividend_start", "เริ่มจ่ายปันผล ปีที่",            4,       ""),
    ("init_invest",    "การลงทุนเริ่มต้นในสินทรัพย์ (THB)", 530000, "สินทรัพย์ + setup"),
    ("depreciation_y1_y3",  "ค่าเสื่อมราคา ปี 1-3 (THB/ปี)",   136000, "ซอฟต์แวร์+ตราสินค้า+อุปกรณ์"),
    ("depreciation_y4_y5",  "ค่าเสื่อมราคา ปี 4-5 (THB/ปี)",   36000,  "ตราสินค้า+อุปกรณ์ (ซอฟต์แวร์หมด)"),
    ("marketing_y1", "Marketing ปีที่ 1 (THB)", 200000, ""),
    ("marketing_y2", "Marketing ปีที่ 2 (THB)", 400000, ""),
    ("marketing_y3", "Marketing ปีที่ 3 (THB)", 500000, ""),
    ("marketing_y4", "Marketing ปีที่ 4 (THB)", 600000, ""),
    ("marketing_y5", "Marketing ปีที่ 5 (THB)", 700000, ""),
    ("other_y1",     "ค่าใช้จ่ายอื่น Y1 (THB)", 120000, "office, admin, travel"),
    ("other_y2",     "ค่าใช้จ่ายอื่น Y2 (THB)", 150000, ""),
    ("other_y3",     "ค่าใช้จ่ายอื่น Y3 (THB)", 180000, ""),
    ("other_y4",     "ค่าใช้จ่ายอื่น Y4 (THB)", 222000, ""),
    ("other_y5",     "ค่าใช้จ่ายอื่น Y5 (THB)", 276000, ""),
    ("salary_mult_y2", "ตัวคูณเงินเดือน Y2", 1.25, ""),
    ("salary_mult_y3", "ตัวคูณเงินเดือน Y3", 2.083, ""),
    ("salary_mult_y4", "ตัวคูณเงินเดือน Y4", 2.75, ""),
    ("salary_mult_y5", "ตัวคูณเงินเดือน Y5", 3.792, ""),
    ("wc_change_y1", "WC change Y1 (THB)", -100000, "ติดลบ=WC เพิ่ม"),
    ("wc_change_y2", "WC change Y2 (THB)", -150000, ""),
    ("wc_change_y3", "WC change Y3 (THB)", -120000, ""),
    ("wc_change_y4", "WC change Y4 (THB)", -100000, ""),
    ("wc_change_y5", "WC change Y5 (THB)", -100000, ""),
    ("capex_y1", "CapEx Y1 (THB)", -50000,  "ติดลบ=จ่าย"),
    ("capex_y2", "CapEx Y2 (THB)", -80000,  ""),
    ("capex_y3", "CapEx Y3 (THB)", -100000, ""),
    ("capex_y4", "CapEx Y4 (THB)", -50000,  ""),
    ("capex_y5", "CapEx Y5 (THB)", -50000,  ""),
]
for key, label, value, note in growth_rows:
    ws.cell(row=row, column=1, value=label).font = TH_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=1).border = BORDER
    if key.startswith(("growth_", "loan_term", "dividend_start")):
        fmt = FMT_INT
    elif "rate" in key or "wacc" in key or "pct" in key or "mult" in key:
        fmt = "0.000"
    else:
        fmt = FMT_THB
    style_cell(ws.cell(row=row, column=2, value=value),
               font=TH_BOLD, fill=INPUT_FILL, fmt=fmt)
    ws.cell(row=row, column=3, value=note).font = TH_SUBTITLE
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="left")
    ws.cell(row=row, column=3).border = BORDER
    A[key] = f"Assumptions!$B${row}"
    row += 1

# ============================================================
# Sheet 2: Loan amortization (5 years)
# ============================================================
loan = wb.create_sheet("Loan")
set_col_widths(loan, [10, 20, 18, 18, 18, 20])

loan.cell(row=1, column=1, value="Loan Amortization — 5 ปี").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
loan.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
row = 3
write_header(loan, row, ["ปี", "เงินต้นคงค้างต้นปี", "ดอกเบี้ย", "เงินต้นที่ชำระ", "ชำระรวมรายปี", "เงินต้นคงค้างสิ้นปี"])
row += 1

# Annual payment: PMT = PV * r * (1+r)^n / ((1+r)^n - 1)
# Using Excel PMT: PMT(rate, nper, pv) but we want positive payment
# Use formula: = -PMT(rate, nper, pv)
loan_start_row = row
for y in range(1, 6):
    loan.cell(row=row, column=1, value=y).font = TH_BOLD
    loan.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    loan.cell(row=row, column=1).border = BORDER
    # Opening balance
    if y == 1:
        bal_formula = f"={A['loan_amount']}"
    else:
        bal_formula = f"=F{row - 1}"
    style_cell(loan.cell(row=row, column=2, value=bal_formula),
               font=TH_FONT, fmt=FMT_THB)
    # Interest = balance * rate
    style_cell(loan.cell(row=row, column=3, value=f"=B{row}*{A['loan_rate']}"),
               font=TH_FONT, fmt=FMT_THB)
    # Annual payment (constant) — PMT formula
    style_cell(loan.cell(row=row, column=5,
                         value=f"=-PMT({A['loan_rate']},{A['loan_term']},{A['loan_amount']})"),
               font=TH_FONT, fmt=FMT_THB)
    # Principal = payment - interest
    style_cell(loan.cell(row=row, column=4, value=f"=E{row}-C{row}"),
               font=TH_FONT, fmt=FMT_THB)
    # Closing balance = opening - principal
    style_cell(loan.cell(row=row, column=6, value=f"=B{row}-D{row}"),
               font=TH_FONT, fill=SUBTOTAL_FILL, fmt=FMT_THB)
    row += 1
loan_end_row = row - 1

# Totals
loan.cell(row=row, column=1, value="รวม").font = TH_BOLD
loan.cell(row=row, column=1).fill = TOTAL_FILL
loan.cell(row=row, column=1).alignment = Alignment(horizontal="center")
loan.cell(row=row, column=1).border = BORDER
for col in ["C", "D", "E"]:
    c = ord(col) - ord("A") + 1
    style_cell(loan.cell(row=row, column=c,
                         value=f"=SUM({col}{loan_start_row}:{col}{loan_end_row})"),
               font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)

LOAN_ROWS = {
    "interest_y1": loan_start_row,      # col C
    "interest_y2": loan_start_row + 1,
    "interest_y3": loan_start_row + 2,
    "interest_y4": loan_start_row + 3,
    "interest_y5": loan_start_row + 4,
    "principal_y1": loan_start_row,     # col D
    "principal_y2": loan_start_row + 1,
    "principal_y3": loan_start_row + 2,
    "principal_y4": loan_start_row + 3,
    "principal_y5": loan_start_row + 4,
    "balance_y1": loan_start_row,       # col F (end of year)
    "balance_y2": loan_start_row + 1,
    "balance_y3": loan_start_row + 2,
    "balance_y4": loan_start_row + 3,
    "balance_y5": loan_start_row + 4,
}

# ============================================================
# Sheet 3: Assets (§6.1) — unchanged
# ============================================================
assets = wb.create_sheet("Assets")
set_col_widths(assets, [35, 40, 15, 12, 18])

assets.cell(row=1, column=1, value="สินทรัพย์ที่ใช้ในการประกอบธุรกิจ").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
assets.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
row = 3
write_header(assets, row, ["รายการ", "รายละเอียด", "มูลค่า (THB)", "อายุ (ปี)", "ค่าเสื่อม/ปี"])
row += 1

asset_items = [
    ("ซอฟต์แวร์ระบบ VIRIYA", "chat-api + ingest + admin + portal", 300000, 3),
    ("ตราสินค้า VIRIYA", "logo + CI + trademark", 30000, 5),
    ("Domain + SSL", "รายจ่ายต่อเนื่อง (OpEx)", 5000, None),
    ("อุปกรณ์สำนักงาน", "คอม + จอ + อุปกรณ์", 150000, 5),
]
asset_start = row
for label, detail, value, life in asset_items:
    assets.cell(row=row, column=1, value=label).font = TH_FONT
    assets.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    assets.cell(row=row, column=1).border = BORDER
    assets.cell(row=row, column=2, value=detail).font = TH_SUBTITLE
    assets.cell(row=row, column=2).alignment = Alignment(horizontal="left", wrap_text=True)
    assets.cell(row=row, column=2).border = BORDER
    style_cell(assets.cell(row=row, column=3, value=value), font=TH_FONT, fmt=FMT_THB)
    if life:
        style_cell(assets.cell(row=row, column=4, value=life), font=TH_FONT, fmt=FMT_INT, align="center")
        style_cell(assets.cell(row=row, column=5, value=f"=C{row}/D{row}"), font=TH_FONT, fmt=FMT_THB)
    else:
        style_cell(assets.cell(row=row, column=4, value="—"), font=TH_SUBTITLE, align="center")
        style_cell(assets.cell(row=row, column=5, value=0), font=TH_SUBTITLE, fmt=FMT_THB)
    row += 1
asset_end = row - 1

# Totals
assets.cell(row=row, column=1, value="รวม").font = TH_BOLD
assets.cell(row=row, column=1).fill = TOTAL_FILL
assets.cell(row=row, column=1).border = BORDER
assets.cell(row=row, column=2).fill = TOTAL_FILL
assets.cell(row=row, column=2).border = BORDER
style_cell(assets.cell(row=row, column=3, value=f"=SUM(C{asset_start}:C{asset_end})"),
           font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)
assets.cell(row=row, column=4).fill = TOTAL_FILL
assets.cell(row=row, column=4).border = BORDER
style_cell(assets.cell(row=row, column=5, value=f"=SUM(E{asset_start}:E{asset_end})"),
           font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)
assets_total_dep_row = row

# ============================================================
# Sheet 4: SG&A (5 years)
# ============================================================
sga = wb.create_sheet("SG&A")
set_col_widths(sga, [45, 14, 14, 14, 14, 14])

sga.cell(row=1, column=1, value="ค่าใช้จ่ายในการขายและการบริหาร (SG&A) 5 ปี").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
sga.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
row = 3
write_header(sga, row, ["รายการ", "ปีที่ 1", "ปีที่ 2", "ปีที่ 3", "ปีที่ 4", "ปีที่ 5"])
row += 1

# Selling
section_banner(sga, row, "ค่าใช้จ่ายในการขาย (Selling)", n_cols=6)
row += 1

selling_start = row
# Let marketing_yN be linked to assumption; ads = 40% of marketing, activities = 30%, content = 30%
sga.cell(row=row, column=1, value="  ค่าโฆษณาออนไลน์").font = TH_FONT
sga.cell(row=row, column=1).alignment = Alignment(horizontal="left")
sga.cell(row=row, column=1).border = BORDER
for i, key in enumerate(["marketing_y1", "marketing_y2", "marketing_y3", "marketing_y4", "marketing_y5"], 2):
    style_cell(sga.cell(row=row, column=i, value=f"={A[key]}*0.4"),
               font=TH_FONT, fmt=FMT_THB)
row += 1

sga.cell(row=row, column=1, value="  กิจกรรมการขายและสาธิต").font = TH_FONT
sga.cell(row=row, column=1).alignment = Alignment(horizontal="left")
sga.cell(row=row, column=1).border = BORDER
for i, key in enumerate(["marketing_y1", "marketing_y2", "marketing_y3", "marketing_y4", "marketing_y5"], 2):
    style_cell(sga.cell(row=row, column=i, value=f"={A[key]}*0.3"),
               font=TH_FONT, fmt=FMT_THB)
row += 1

sga.cell(row=row, column=1, value="  เนื้อหาการตลาด").font = TH_FONT
sga.cell(row=row, column=1).alignment = Alignment(horizontal="left")
sga.cell(row=row, column=1).border = BORDER
for i, key in enumerate(["marketing_y1", "marketing_y2", "marketing_y3", "marketing_y4", "marketing_y5"], 2):
    style_cell(sga.cell(row=row, column=i, value=f"={A[key]}*0.3"),
               font=TH_FONT, fmt=FMT_THB)
row += 1
selling_end = row - 1

sga.cell(row=row, column=1, value="รวมค่าใช้จ่ายในการขาย").font = TH_BOLD
sga.cell(row=row, column=1).fill = SUBTOTAL_FILL
sga.cell(row=row, column=1).alignment = Alignment(horizontal="left")
sga.cell(row=row, column=1).border = BORDER
for i, col in enumerate(["B", "C", "D", "E", "F"], 2):
    style_cell(sga.cell(row=row, column=i, value=f"=SUM({col}{selling_start}:{col}{selling_end})"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
selling_total_row = row
row += 2

# Admin
section_banner(sga, row, "ค่าใช้จ่ายในการบริหาร (G&A)", n_cols=6)
row += 1

admin_start = row
# Salary = salary_total * 12 * multiplier
sga.cell(row=row, column=1, value="  ค่าแรงบุคลากร").font = TH_FONT
sga.cell(row=row, column=1).alignment = Alignment(horizontal="left")
sga.cell(row=row, column=1).border = BORDER
style_cell(sga.cell(row=row, column=2, value=f"={A['salary_total']}*12"),
           font=TH_FONT, fmt=FMT_THB)
style_cell(sga.cell(row=row, column=3, value=f"={A['salary_total']}*12*{A['salary_mult_y2']}"),
           font=TH_FONT, fmt=FMT_THB)
style_cell(sga.cell(row=row, column=4, value=f"={A['salary_total']}*12*{A['salary_mult_y3']}"),
           font=TH_FONT, fmt=FMT_THB)
style_cell(sga.cell(row=row, column=5, value=f"={A['salary_total']}*12*{A['salary_mult_y4']}"),
           font=TH_FONT, fmt=FMT_THB)
style_cell(sga.cell(row=row, column=6, value=f"={A['salary_total']}*12*{A['salary_mult_y5']}"),
           font=TH_FONT, fmt=FMT_THB)
row += 1

# Other admin items — link to other_yN (which is aggregate of rent/travel/utilities/legal)
sga.cell(row=row, column=1, value="  ค่าเช่า + ค่าเดินทาง + สาธารณูปโภค + บัญชี/กฎหมาย").font = TH_FONT
sga.cell(row=row, column=1).alignment = Alignment(horizontal="left")
sga.cell(row=row, column=1).border = BORDER
for i, key in enumerate(["other_y1", "other_y2", "other_y3", "other_y4", "other_y5"], 2):
    style_cell(sga.cell(row=row, column=i, value=f"={A[key]}"),
               font=TH_FONT, fmt=FMT_THB)
row += 1
admin_end = row - 1

sga.cell(row=row, column=1, value="รวมค่าใช้จ่ายในการบริหาร").font = TH_BOLD
sga.cell(row=row, column=1).fill = SUBTOTAL_FILL
sga.cell(row=row, column=1).alignment = Alignment(horizontal="left")
sga.cell(row=row, column=1).border = BORDER
for i, col in enumerate(["B", "C", "D", "E", "F"], 2):
    style_cell(sga.cell(row=row, column=i, value=f"=SUM({col}{admin_start}:{col}{admin_end})"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
admin_total_row = row
row += 2

# Grand total
sga.cell(row=row, column=1, value="รวม SG&A").font = TH_BOLD
sga.cell(row=row, column=1).fill = TOTAL_FILL
sga.cell(row=row, column=1).alignment = Alignment(horizontal="left")
sga.cell(row=row, column=1).border = BORDER
for i, col in enumerate(["B", "C", "D", "E", "F"], 2):
    style_cell(sga.cell(row=row, column=i,
                        value=f"={col}{selling_total_row}+{col}{admin_total_row}"),
               font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)
sga_total_row = row

# ============================================================
# Sheet 5: P&L (5 years)
# ============================================================
pl = wb.create_sheet("P&L")
set_col_widths(pl, [42, 14, 14, 14, 14, 14])

pl.cell(row=1, column=1, value="งบกำไรขาดทุน 5 ปี").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
pl.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
row = 3
write_header(pl, row, ["รายการ", "ปีที่ 1", "ปีที่ 2", "ปีที่ 3", "ปีที่ 4", "ปีที่ 5"])
row += 1

# Tenant counts
write_label_row(pl, row, "Tenant ณ สิ้นปี",
                [f"={A['growth_y1']}*12",
                 f"=B{row}+{A['growth_y2']}*12",
                 f"=C{row}+{A['growth_y3']}*12",
                 f"=D{row}+{A['growth_y4']}*12",
                 f"=E{row}+{A['growth_y5']}*12"],
                font=TH_BOLD, fmt=FMT_INT)
end_row = row
row += 1

write_label_row(pl, row, "Tenant เฉลี่ย (start + growth×6.5)",
                [f"=0+{A['growth_y1']}*6.5",
                 f"=B{end_row}+{A['growth_y2']}*6.5",
                 f"=C{end_row}+{A['growth_y3']}*6.5",
                 f"=D{end_row}+{A['growth_y4']}*6.5",
                 f"=E{end_row}+{A['growth_y5']}*6.5"],
                font=TH_BOLD, fmt=FMT_INT)
avg_row = row
row += 2

# Revenue
section_banner(pl, row, "รายได้", n_cols=6)
row += 1

write_label_row(pl, row, "  ค่าสมาชิก",
                [f"=B{avg_row}*{A['avg_revenue']}*12",
                 f"=C{avg_row}*{A['avg_revenue']}*12",
                 f"=D{avg_row}*{A['avg_revenue']}*12",
                 f"=E{avg_row}*{A['avg_revenue']}*12",
                 f"=F{avg_row}*{A['avg_revenue']}*12"])
sub_rev_row = row
row += 1

write_label_row(pl, row, "  Setup fee",
                [f"={A['growth_y1']}*12*{A['setup_fee']}",
                 f"={A['growth_y2']}*12*{A['setup_fee']}",
                 f"={A['growth_y3']}*12*{A['setup_fee']}",
                 f"={A['growth_y4']}*12*{A['setup_fee']}",
                 f"={A['growth_y5']}*12*{A['setup_fee']}"])
setup_row = row
row += 1

write_label_row(pl, row, "รวมรายได้",
                [f"=B{sub_rev_row}+B{setup_row}",
                 f"=C{sub_rev_row}+C{setup_row}",
                 f"=D{sub_rev_row}+D{setup_row}",
                 f"=E{sub_rev_row}+E{setup_row}",
                 f"=F{sub_rev_row}+F{setup_row}"],
                font=TH_BOLD, fill=SUBTOTAL_FILL)
total_rev_row = row
row += 2

# Variable costs
section_banner(pl, row, "ต้นทุนผันแปร", n_cols=6)
row += 1

write_label_row(pl, row, "  AI API + Pinecone",
                [f"=B{avg_row}*{A['var_total']}*12",
                 f"=C{avg_row}*{A['var_total']}*12",
                 f"=D{avg_row}*{A['var_total']}*12",
                 f"=E{avg_row}*{A['var_total']}*12",
                 f"=F{avg_row}*{A['var_total']}*12"])
api_cost_row = row
row += 1

write_label_row(pl, row, "  Onboard tenant ใหม่",
                [f"={A['growth_y1']}*12*{A['onboard_cost']}",
                 f"={A['growth_y2']}*12*{A['onboard_cost']}",
                 f"={A['growth_y3']}*12*{A['onboard_cost']}",
                 f"={A['growth_y4']}*12*{A['onboard_cost']}",
                 f"={A['growth_y5']}*12*{A['onboard_cost']}"])
onboard_row = row
row += 1

write_label_row(pl, row, "รวมต้นทุนผันแปร",
                [f"=B{api_cost_row}+B{onboard_row}",
                 f"=C{api_cost_row}+C{onboard_row}",
                 f"=D{api_cost_row}+D{onboard_row}",
                 f"=E{api_cost_row}+E{onboard_row}",
                 f"=F{api_cost_row}+F{onboard_row}"],
                font=TH_BOLD, fill=SUBTOTAL_FILL)
total_var_row = row
row += 2

# Gross Profit
write_label_row(pl, row, "กำไรขั้นต้น (Gross Profit)",
                [f"=B{total_rev_row}-B{total_var_row}",
                 f"=C{total_rev_row}-C{total_var_row}",
                 f"=D{total_rev_row}-D{total_var_row}",
                 f"=E{total_rev_row}-E{total_var_row}",
                 f"=F{total_rev_row}-F{total_var_row}"],
                font=TH_BOLD, fill=TOTAL_FILL)
gross_row = row
row += 1

write_label_row(pl, row, "  อัตรากำไรขั้นต้น (%)",
                [f"=B{gross_row}/B{total_rev_row}",
                 f"=C{gross_row}/C{total_rev_row}",
                 f"=D{gross_row}/D{total_rev_row}",
                 f"=E{gross_row}/E{total_rev_row}",
                 f"=F{gross_row}/F{total_rev_row}"],
                fmt=FMT_PCT)
row += 2

# OpEx
section_banner(pl, row, "ค่าใช้จ่ายในการดำเนินงาน (SG&A + infra)", n_cols=6)
row += 1

write_label_row(pl, row, "  SG&A (link SG&A sheet)",
                [f"='SG&A'!B{sga_total_row}",
                 f"='SG&A'!C{sga_total_row}",
                 f"='SG&A'!D{sga_total_row}",
                 f"='SG&A'!E{sga_total_row}",
                 f"='SG&A'!F{sga_total_row}"])
opex_sga_row = row
row += 1

write_label_row(pl, row, "  โครงสร้างพื้นฐานระบบ",
                [f"={A['fix_total']}*12"] * 5)
opex_infra_row = row
row += 1

write_label_row(pl, row, "รวม OpEx",
                [f"=B{opex_sga_row}+B{opex_infra_row}",
                 f"=C{opex_sga_row}+C{opex_infra_row}",
                 f"=D{opex_sga_row}+D{opex_infra_row}",
                 f"=E{opex_sga_row}+E{opex_infra_row}",
                 f"=F{opex_sga_row}+F{opex_infra_row}"],
                font=TH_BOLD, fill=SUBTOTAL_FILL)
total_opex_row = row
row += 2

# EBITDA (gross - opex)
write_label_row(pl, row, "EBITDA (Gross - OpEx)",
                [f"=B{gross_row}-B{total_opex_row}",
                 f"=C{gross_row}-C{total_opex_row}",
                 f"=D{gross_row}-D{total_opex_row}",
                 f"=E{gross_row}-E{total_opex_row}",
                 f"=F{gross_row}-F{total_opex_row}"],
                font=TH_BOLD, fill=TOTAL_FILL)
ebitda_row = row
row += 1

# Depreciation (Y1-Y3: 136K, Y4-Y5: 36K)
write_label_row(pl, row, "ค่าเสื่อมราคา",
                [f"={A['depreciation_y1_y3']}",
                 f"={A['depreciation_y1_y3']}",
                 f"={A['depreciation_y1_y3']}",
                 f"={A['depreciation_y4_y5']}",
                 f"={A['depreciation_y4_y5']}"])
dep_row = row
row += 1

# EBIT
write_label_row(pl, row, "EBIT (EBITDA - Dep)",
                [f"=B{ebitda_row}-B{dep_row}",
                 f"=C{ebitda_row}-C{dep_row}",
                 f"=D{ebitda_row}-D{dep_row}",
                 f"=E{ebitda_row}-E{dep_row}",
                 f"=F{ebitda_row}-F{dep_row}"],
                font=TH_BOLD, fill=TOTAL_FILL)
ebit_row = row
row += 1

# Interest from Loan sheet
write_label_row(pl, row, "ดอกเบี้ยจ่าย (จาก Loan sheet)",
                [f"=Loan!C{LOAN_ROWS['interest_y1']}",
                 f"=Loan!C{LOAN_ROWS['interest_y2']}",
                 f"=Loan!C{LOAN_ROWS['interest_y3']}",
                 f"=Loan!C{LOAN_ROWS['interest_y4']}",
                 f"=Loan!C{LOAN_ROWS['interest_y5']}"])
interest_row = row
row += 1

# EBT
write_label_row(pl, row, "กำไรก่อนหักภาษี (EBT)",
                [f"=B{ebit_row}-B{interest_row}",
                 f"=C{ebit_row}-C{interest_row}",
                 f"=D{ebit_row}-D{interest_row}",
                 f"=E{ebit_row}-E{interest_row}",
                 f"=F{ebit_row}-F{interest_row}"],
                font=TH_BOLD, fill=TOTAL_FILL)
ebt_row = row
row += 1

# Tax — only on positive EBT
write_label_row(pl, row, "ภาษีเงินได้ (20% ของ EBT ถ้า +)",
                [f"=MAX(0, B{ebt_row}*{A['tax_rate']})",
                 f"=MAX(0, C{ebt_row}*{A['tax_rate']})",
                 f"=MAX(0, D{ebt_row}*{A['tax_rate']})",
                 f"=MAX(0, E{ebt_row}*{A['tax_rate']})",
                 f"=MAX(0, F{ebt_row}*{A['tax_rate']})"])
tax_row = row
row += 1

# Net profit
write_label_row(pl, row, "กำไรสุทธิ (Net Profit)",
                [f"=B{ebt_row}-B{tax_row}",
                 f"=C{ebt_row}-C{tax_row}",
                 f"=D{ebt_row}-D{tax_row}",
                 f"=E{ebt_row}-E{tax_row}",
                 f"=F{ebt_row}-F{tax_row}"],
                font=TH_BOLD, fill=TOTAL_FILL)
net_row = row
row += 1

write_label_row(pl, row, "  อัตรากำไรสุทธิ (%)",
                [f"=B{net_row}/B{total_rev_row}",
                 f"=C{net_row}/C{total_rev_row}",
                 f"=D{net_row}/D{total_rev_row}",
                 f"=E{net_row}/E{total_rev_row}",
                 f"=F{net_row}/F{total_rev_row}"],
                fmt=FMT_PCT)
row += 2

# Dividend (only from Y4)
write_label_row(pl, row, "การจ่ายปันผล (30% จาก Y4)",
                [f"=IF(1>={A['dividend_start']},MAX(0,B{net_row}*{A['dividend_pct']}),0)",
                 f"=IF(2>={A['dividend_start']},MAX(0,C{net_row}*{A['dividend_pct']}),0)",
                 f"=IF(3>={A['dividend_start']},MAX(0,D{net_row}*{A['dividend_pct']}),0)",
                 f"=IF(4>={A['dividend_start']},MAX(0,E{net_row}*{A['dividend_pct']}),0)",
                 f"=IF(5>={A['dividend_start']},MAX(0,F{net_row}*{A['dividend_pct']}),0)"])
div_row = row
row += 1

PL = {
    "net_profit": net_row,
    "total_revenue": total_rev_row,
    "gross_profit": gross_row,
    "ebitda": ebitda_row,
    "ebit": ebit_row,
    "ebt": ebt_row,
    "dep": dep_row,
    "tax": tax_row,
    "interest": interest_row,
    "dividend": div_row,
}

# ============================================================
# Sheet 6: CashFlow (5 years, cols B-G = Y0..Y5)
# ============================================================
cf = wb.create_sheet("CashFlow")
set_col_widths(cf, [42, 14, 14, 14, 14, 14, 14])

cf.cell(row=1, column=1, value="งบกระแสเงินสด 5 ปี").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
cf.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
row = 3
write_header(cf, row, ["รายการ", "ปีที่ 0", "ปีที่ 1", "ปีที่ 2", "ปีที่ 3", "ปีที่ 4", "ปีที่ 5"])
row += 1

# Operating
section_banner(cf, row, "กระแสเงินสดจากการดำเนินงาน", n_cols=7)
row += 1

# Net profit (Y0=0, Y1-Y5 from PL)
cf.cell(row=row, column=1, value="  กำไรสุทธิ").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=0), font=TH_FONT, fmt=FMT_THB)
for i, col_src in enumerate(["B", "C", "D", "E", "F"], 3):
    style_cell(cf.cell(row=row, column=i, value=f"='P&L'!{col_src}{PL['net_profit']}"),
               font=TH_FONT, fmt=FMT_THB)
np_cf_row = row
row += 1

# Add depreciation
cf.cell(row=row, column=1, value="  บวก: ค่าเสื่อมราคา").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=0), font=TH_FONT, fmt=FMT_THB)
for i, col_src in enumerate(["B", "C", "D", "E", "F"], 3):
    style_cell(cf.cell(row=row, column=i, value=f"='P&L'!{col_src}{PL['dep']}"),
               font=TH_FONT, fmt=FMT_THB)
dep_cf_row = row
row += 1

# WC change
cf.cell(row=row, column=1, value="  การเปลี่ยนแปลงทุนหมุนเวียน").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=0), font=TH_FONT, fmt=FMT_THB)
for i, key in enumerate(["wc_change_y1", "wc_change_y2", "wc_change_y3", "wc_change_y4", "wc_change_y5"], 3):
    style_cell(cf.cell(row=row, column=i, value=f"={A[key]}"),
               font=TH_FONT, fmt=FMT_THB)
wc_cf_row = row
row += 1

# Total operating CF
cf.cell(row=row, column=1, value="รวมกระแสเงินสดจากการดำเนินงาน").font = TH_BOLD
cf.cell(row=row, column=1).fill = SUBTOTAL_FILL
cf.cell(row=row, column=1).border = BORDER
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(cf.cell(row=row, column=i,
                       value=f"={col}{np_cf_row}+{col}{dep_cf_row}+{col}{wc_cf_row}"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
ocf_row = row
row += 2

# Investing
section_banner(cf, row, "กระแสเงินสดจากการลงทุน", n_cols=7)
row += 1

cf.cell(row=row, column=1, value="  การลงทุนเริ่มต้น (สินทรัพย์)").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=f"=-{A['init_invest']}"), font=TH_FONT, fmt=FMT_THB)
for i in range(3, 8):
    style_cell(cf.cell(row=row, column=i, value=0), font=TH_FONT, fmt=FMT_THB)
init_row = row
row += 1

cf.cell(row=row, column=1, value="  CapEx เพิ่มเติม").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=0), font=TH_FONT, fmt=FMT_THB)
for i, key in enumerate(["capex_y1", "capex_y2", "capex_y3", "capex_y4", "capex_y5"], 3):
    style_cell(cf.cell(row=row, column=i, value=f"={A[key]}"),
               font=TH_FONT, fmt=FMT_THB)
capex_cf_row = row
row += 1

cf.cell(row=row, column=1, value="รวมกระแสเงินสดจากการลงทุน").font = TH_BOLD
cf.cell(row=row, column=1).fill = SUBTOTAL_FILL
cf.cell(row=row, column=1).border = BORDER
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(cf.cell(row=row, column=i, value=f"={col}{init_row}+{col}{capex_cf_row}"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
icf_row = row
row += 2

# Financing
section_banner(cf, row, "กระแสเงินสดจากกิจกรรมทางการเงิน", n_cols=7)
row += 1

# Equity (Y0)
cf.cell(row=row, column=1, value="  เงินทุนผู้ถือหุ้น").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=f"={A['founder_cap']}"), font=TH_FONT, fmt=FMT_THB)
for i in range(3, 8):
    style_cell(cf.cell(row=row, column=i, value=0), font=TH_FONT, fmt=FMT_THB)
equity_row = row
row += 1

# Loan receipt (Y0)
cf.cell(row=row, column=1, value="  เงินกู้ระยะยาว (received)").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=f"={A['loan_amount']}"), font=TH_FONT, fmt=FMT_THB)
for i in range(3, 8):
    style_cell(cf.cell(row=row, column=i, value=0), font=TH_FONT, fmt=FMT_THB)
loan_recv_row = row
row += 1

# Loan principal payment (Y1-Y5)
cf.cell(row=row, column=1, value="  การชำระเงินต้น").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=0), font=TH_FONT, fmt=FMT_THB)
for i, y_key in enumerate(["principal_y1", "principal_y2", "principal_y3", "principal_y4", "principal_y5"], 3):
    style_cell(cf.cell(row=row, column=i, value=f"=-Loan!D{LOAN_ROWS[y_key]}"),
               font=TH_FONT, fmt=FMT_THB)
principal_row = row
row += 1

# Dividend payments
cf.cell(row=row, column=1, value="  การจ่ายปันผล").font = TH_FONT
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
cf.cell(row=row, column=1).border = BORDER
style_cell(cf.cell(row=row, column=2, value=0), font=TH_FONT, fmt=FMT_THB)
for i, col_src in enumerate(["B", "C", "D", "E", "F"], 3):
    style_cell(cf.cell(row=row, column=i, value=f"=-'P&L'!{col_src}{PL['dividend']}"),
               font=TH_FONT, fmt=FMT_THB)
div_cf_row = row
row += 1

cf.cell(row=row, column=1, value="รวมกระแสเงินสดจากกิจกรรมทางการเงิน").font = TH_BOLD
cf.cell(row=row, column=1).fill = SUBTOTAL_FILL
cf.cell(row=row, column=1).border = BORDER
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(cf.cell(row=row, column=i,
                       value=f"={col}{equity_row}+{col}{loan_recv_row}+{col}{principal_row}+{col}{div_cf_row}"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
fcf_row = row
row += 2

# Net change
cf.cell(row=row, column=1, value="กระแสเงินสดสุทธิเปลี่ยนแปลง").font = TH_BOLD
cf.cell(row=row, column=1).fill = TOTAL_FILL
cf.cell(row=row, column=1).border = BORDER
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(cf.cell(row=row, column=i,
                       value=f"={col}{ocf_row}+{col}{icf_row}+{col}{fcf_row}"),
               font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)
net_change_row = row
row += 1

# Ending cash cumulative
cf.cell(row=row, column=1, value="เงินสดคงเหลือสะสม ณ สิ้นงวด").font = TH_BOLD
cf.cell(row=row, column=1).fill = TOTAL_FILL
cf.cell(row=row, column=1).border = BORDER
cf.cell(row=row, column=1).alignment = Alignment(horizontal="left")
# Y0 = net_change Y0
style_cell(cf.cell(row=row, column=2, value=f"=B{net_change_row}"),
           font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)
for i, (col, prev) in enumerate([("C", "B"), ("D", "C"), ("E", "D"), ("F", "E"), ("G", "F")], 3):
    style_cell(cf.cell(row=row, column=i, value=f"={prev}{row}+{col}{net_change_row}"),
               font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)
ending_cash_row = row
row += 1

CF = {
    "ocf": ocf_row,
    "icf": icf_row,
    "fcf": fcf_row,
    "net_change": net_change_row,
    "ending_cash": ending_cash_row,
}

# ============================================================
# Sheet 7: BalanceSheet (5 years)
# ============================================================
bs = wb.create_sheet("BalanceSheet")
set_col_widths(bs, [42, 14, 14, 14, 14, 14, 14])

bs.cell(row=1, column=1, value="งบแสดงฐานะทางการเงิน 5 ปี").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
bs.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
bs.cell(row=2, column=1, value="กำไรสะสมคำนวณเป็น plug เพื่อให้ Balance Sheet สมดุล").font = TH_SUBTITLE
bs.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
row = 4
write_header(bs, row, ["รายการ", "ปีที่ 0", "ปีที่ 1", "ปีที่ 2", "ปีที่ 3", "ปีที่ 4", "ปีที่ 5"])
row += 1

# Current assets
section_banner(bs, row, "สินทรัพย์หมุนเวียน", n_cols=7)
row += 1

# Cash from CashFlow
bs.cell(row=row, column=1, value="  เงินสด").font = TH_FONT
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
bs.cell(row=row, column=1).border = BORDER
for i, col_src in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"=CashFlow!{col_src}{CF['ending_cash']}"),
               font=TH_FONT, fmt=FMT_THB)
bs_cash_row = row
row += 1

# AR (input)
bs.cell(row=row, column=1, value="  ลูกหนี้การค้า (input)").font = TH_FONT
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
bs.cell(row=row, column=1).border = BORDER
ar_values = [0, 80000, 200000, 280000, 340000, 400000]
for i, v in enumerate(ar_values, 2):
    style_cell(bs.cell(row=row, column=i, value=v),
               font=TH_FONT, fill=INPUT_FILL, fmt=FMT_THB)
bs_ar_row = row
row += 1

# Sub total
bs.cell(row=row, column=1, value="รวมสินทรัพย์หมุนเวียน").font = TH_BOLD
bs.cell(row=row, column=1).fill = SUBTOTAL_FILL
bs.cell(row=row, column=1).border = BORDER
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"={col}{bs_cash_row}+{col}{bs_ar_row}"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
bs_current_row = row
row += 2

# Non-current (input for now)
section_banner(bs, row, "สินทรัพย์ไม่หมุนเวียน", n_cols=7)
row += 1

non_current = [
    ("  อุปกรณ์ (net)", [150000, 170000, 220000, 290000, 310000, 330000]),
    ("  ซอฟต์แวร์ (net)", [300000, 200000, 100000, 0, 0, 0]),
    ("  ตราสินค้า (net)", [30000, 24000, 18000, 12000, 6000, 0]),
]
nc_start = row
for label, vals in non_current:
    bs.cell(row=row, column=1, value=label).font = TH_FONT
    bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    bs.cell(row=row, column=1).border = BORDER
    for i, v in enumerate(vals, 2):
        style_cell(bs.cell(row=row, column=i, value=v),
                   font=TH_FONT, fill=INPUT_FILL, fmt=FMT_THB)
    row += 1
nc_end = row - 1

bs.cell(row=row, column=1, value="รวมสินทรัพย์ไม่หมุนเวียน").font = TH_BOLD
bs.cell(row=row, column=1).fill = SUBTOTAL_FILL
bs.cell(row=row, column=1).border = BORDER
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"=SUM({col}{nc_start}:{col}{nc_end})"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
bs_nc_row = row
row += 2

# Total assets
bs.cell(row=row, column=1, value="รวมสินทรัพย์").font = TH_BOLD
bs.cell(row=row, column=1).fill = TOTAL_FILL
bs.cell(row=row, column=1).border = BORDER
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"={col}{bs_current_row}+{col}{bs_nc_row}"),
               font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)
bs_total_assets_row = row
row += 2

# Current liab
section_banner(bs, row, "หนี้สินหมุนเวียน", n_cols=7)
row += 1

liab_items = [
    ("  เจ้าหนี้การค้า (input)", [0, 30000, 60000, 80000, 100000, 120000], True),
    ("  ค่าใช้จ่ายค้างจ่าย (input)", [0, 20000, 40000, 60000, 80000, 100000], True),
]
cliab_start = row
for label, vals, is_input in liab_items:
    bs.cell(row=row, column=1, value=label).font = TH_FONT
    bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    bs.cell(row=row, column=1).border = BORDER
    fill = INPUT_FILL if is_input else None
    for i, v in enumerate(vals, 2):
        style_cell(bs.cell(row=row, column=i, value=v),
                   font=TH_FONT, fill=fill, fmt=FMT_THB)
    row += 1

# Tax payable (link to P&L)
bs.cell(row=row, column=1, value="  ภาษีค้างจ่าย (link P&L)").font = TH_FONT
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
bs.cell(row=row, column=1).border = BORDER
style_cell(bs.cell(row=row, column=2, value=0), font=TH_FONT, fmt=FMT_THB)
for i, col_src in enumerate(["B", "C", "D", "E", "F"], 3):
    style_cell(bs.cell(row=row, column=i, value=f"='P&L'!{col_src}{PL['tax']}"),
               font=TH_FONT, fmt=FMT_THB)
row += 1
cliab_end = row - 1

bs.cell(row=row, column=1, value="รวมหนี้สินหมุนเวียน").font = TH_BOLD
bs.cell(row=row, column=1).fill = SUBTOTAL_FILL
bs.cell(row=row, column=1).border = BORDER
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"=SUM({col}{cliab_start}:{col}{cliab_end})"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
bs_cliab_row = row
row += 2

# Long-term debt
section_banner(bs, row, "หนี้สินไม่หมุนเวียน", n_cols=7)
row += 1

bs.cell(row=row, column=1, value="  เงินกู้ระยะยาว (จาก Loan sheet)").font = TH_FONT
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
bs.cell(row=row, column=1).border = BORDER
# Y0 = full loan, Y1-Y5 = balance end of year
style_cell(bs.cell(row=row, column=2, value=f"={A['loan_amount']}"), font=TH_FONT, fmt=FMT_THB)
for i, y_key in enumerate(["balance_y1", "balance_y2", "balance_y3", "balance_y4", "balance_y5"], 3):
    style_cell(bs.cell(row=row, column=i, value=f"=Loan!F{LOAN_ROWS[y_key]}"),
               font=TH_FONT, fmt=FMT_THB)
bs_ltd_row = row
row += 1

# Total liab
bs.cell(row=row, column=1, value="รวมหนี้สิน").font = TH_BOLD
bs.cell(row=row, column=1).fill = SUBTOTAL_FILL
bs.cell(row=row, column=1).border = BORDER
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"={col}{bs_cliab_row}+{col}{bs_ltd_row}"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
bs_total_liab_row = row
row += 2

# Equity
section_banner(bs, row, "ส่วนของผู้ถือหุ้น", n_cols=7)
row += 1

bs.cell(row=row, column=1, value="  ทุนจดทะเบียน").font = TH_FONT
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
bs.cell(row=row, column=1).border = BORDER
for i in range(2, 8):
    style_cell(bs.cell(row=row, column=i, value=f"={A['founder_cap']}"), font=TH_FONT, fmt=FMT_THB)
bs_paidin_row = row
row += 1

# Retained earnings — plug (Total assets - Total liab - Paid-in)
bs.cell(row=row, column=1, value="  กำไรสะสม (plug)").font = TH_FONT
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
bs.cell(row=row, column=1).border = BORDER
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"={col}{bs_total_assets_row}-{col}{bs_total_liab_row}-{col}{bs_paidin_row}"),
               font=TH_FONT, fmt=FMT_THB)
bs_re_row = row
row += 1

bs.cell(row=row, column=1, value="รวมส่วนของผู้ถือหุ้น").font = TH_BOLD
bs.cell(row=row, column=1).fill = SUBTOTAL_FILL
bs.cell(row=row, column=1).border = BORDER
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"={col}{bs_paidin_row}+{col}{bs_re_row}"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
bs_eq_row = row
row += 2

# Grand total
bs.cell(row=row, column=1, value="รวมหนี้สิน + ส่วนของผู้ถือหุ้น").font = TH_BOLD
bs.cell(row=row, column=1).fill = TOTAL_FILL
bs.cell(row=row, column=1).border = BORDER
bs.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for i, col in enumerate(["B", "C", "D", "E", "F", "G"], 2):
    style_cell(bs.cell(row=row, column=i,
                       value=f"={col}{bs_total_liab_row}+{col}{bs_eq_row}"),
               font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_THB)

# ============================================================
# Sheet 8: BreakEven (unchanged structure)
# ============================================================
be = wb.create_sheet("BreakEven")
set_col_widths(be, [50, 18, 40])
be.cell(row=1, column=1, value="การวิเคราะห์จุดคุ้มทุน").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
be.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
row = 3
write_header(be, row, ["รายการ", "ค่า", "คำอธิบาย"])
row += 1

be_items = [
    ("Contribution margin/tenant/mo", f"={A['avg_revenue']}-{A['var_total']}", "รายได้ - ต้นทุนผันแปร", FMT_THB),
    ("ต้นทุนคงที่/mo (system + salary Y1)", f"={A['fix_total']}+{A['salary_total']}", "", FMT_THB),
    ("จำนวน tenant ที่จุดคุ้มทุน", f"=ROUNDUP(B{row+1}/B{row},0)", "fixed / contribution", FMT_INT),
    ("เดือนถึงจุดคุ้มทุน (Y1 growth)", f"=ROUNDUP(B{row+2}/{A['growth_y1']},0)", "be / growth", FMT_INT),
]
for label, formula, note, fmt in be_items:
    be.cell(row=row, column=1, value=label).font = TH_BOLD
    be.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    be.cell(row=row, column=1).border = BORDER
    style_cell(be.cell(row=row, column=2, value=formula), font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=fmt)
    be.cell(row=row, column=3, value=note).font = TH_SUBTITLE
    be.cell(row=row, column=3).alignment = Alignment(horizontal="left")
    be.cell(row=row, column=3).border = BORDER
    row += 1
be_tenant_row = row - 2

# ============================================================
# Sheet 9: Sensitivity (3 scenarios)
# ============================================================
sens = wb.create_sheet("Sensitivity")
set_col_widths(sens, [45, 14, 14, 14])
sens.cell(row=1, column=1, value="การวิเคราะห์ความอ่อนไหว").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
sens.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
row = 3
write_header(sens, row, ["ตัวแปร", "Worst", "Base", "Best"])
row += 1

scenarios = [
    ("Tenant ใหม่/mo", 1, 2, 3, FMT_INT),
    ("สัดส่วน Starter", 0.80, 0.60, 0.40, FMT_PCT),
    ("สัดส่วน Pro", 0.15, 0.30, 0.40, FMT_PCT),
    ("สัดส่วน Enterprise", 0.05, 0.10, 0.20, FMT_PCT),
]
sens_start_row = row
for label, w, b, bst, fmt in scenarios:
    sens.cell(row=row, column=1, value=label).font = TH_FONT
    sens.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    sens.cell(row=row, column=1).border = BORDER
    for col, v in enumerate([w, b, bst], 2):
        style_cell(sens.cell(row=row, column=col, value=v),
                   font=TH_BOLD, fill=INPUT_FILL, fmt=fmt)
    row += 1

gr_row = sens_start_row
mix_st = sens_start_row + 1
mix_pr = sens_start_row + 2
mix_en = sens_start_row + 3
row += 1

sens.cell(row=row, column=1, value="รายได้เฉลี่ย/tenant/mo").font = TH_BOLD
sens.cell(row=row, column=1).fill = SUBTOTAL_FILL
sens.cell(row=row, column=1).border = BORDER
sens.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for col in ["B", "C", "D"]:
    c = ord(col) - ord("A") + 1
    style_cell(sens.cell(row=row, column=c,
                         value=f"={col}{mix_st}*{A['price_starter']}+{col}{mix_pr}*{A['price_pro']}+{col}{mix_en}*{A['price_enterprise']}"),
               font=TH_BOLD, fill=SUBTOTAL_FILL, fmt=FMT_THB)
row += 1

sens.cell(row=row, column=1, value="Tenant ณ สิ้นปี 5").font = TH_BOLD
sens.cell(row=row, column=1).fill = TOTAL_FILL
sens.cell(row=row, column=1).border = BORDER
sens.cell(row=row, column=1).alignment = Alignment(horizontal="left")
for col in ["B", "C", "D"]:
    c = ord(col) - ord("A") + 1
    style_cell(sens.cell(row=row, column=c, value=f"={col}{gr_row}*60"),  # approximate
               font=TH_BOLD, fill=TOTAL_FILL, fmt=FMT_INT)

# ============================================================
# Sheet 10: Summary (NPV/IRR/MIRR/Payback)
# ============================================================
summary = wb.create_sheet("Summary")
set_col_widths(summary, [48, 18, 45])

summary.cell(row=1, column=1, value="บทสรุปทางการเงิน (NPV/IRR/MIRR/Payback)").font = Font(
    name="TH Sarabun New", size=20, bold=True, color="305496"
)
summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
row = 3

# Project cash flows (at-project level, not with financing)
# Y0 = -founder - loan_receipt + loan_receipt + init_invest = -(founder + init_invest) — but we added loan as financing
# For NPV: use operating CF + investing CF (exclude financing)
write_header(summary, row, ["Project Cash Flows for NPV/IRR", "Y0-Y5", ""])
row += 1

# Build project NCF row (Y0-Y5): OCF + ICF (exclude financing)
# Let's compute: Y0 NCF = - (founder + loan) + (founder + loan - invest) = -invest, but treat Y0 as initial investment outflow only
# For project NPV, Y0 = -3,000,000 (full initial funding), then Y1-Y5 = OCF + ICF
summary.cell(row=row, column=1, value="  Y0 Initial Investment (equity + loan)").font = TH_FONT
summary.cell(row=row, column=1).alignment = Alignment(horizontal="left")
summary.cell(row=row, column=1).border = BORDER
style_cell(summary.cell(row=row, column=2,
                        value=f"=-({A['founder_cap']}+{A['loan_amount']})"),
           font=TH_FONT, fmt=FMT_THB)
ncf_y0_row = row
row += 1

year_labels = ["Y1", "Y2", "Y3", "Y4", "Y5"]
ncf_rows = {}
for y_idx, y_label in enumerate(year_labels, 1):
    summary.cell(row=row, column=1, value=f"  {y_label} NCF = OCF + ICF").font = TH_FONT
    summary.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    summary.cell(row=row, column=1).border = BORDER
    col = ["C", "D", "E", "F", "G"][y_idx - 1]
    style_cell(summary.cell(row=row, column=2,
                            value=f"=CashFlow!{col}{CF['ocf']}+CashFlow!{col}{CF['icf']}"),
               font=TH_FONT, fmt=FMT_THB)
    ncf_rows[y_label] = row
    row += 1

row += 1

# NPV, IRR, MIRR, Payback
write_header(summary, row, ["ตัวชี้วัด", "ค่า", "การประเมิน"])
row += 1

kpi_items = [
    ("NPV (WACC 12%)",
     f"=B{ncf_y0_row}+NPV({A['wacc']},B{ncf_rows['Y1']}:B{ncf_rows['Y5']})",
     FMT_THB,
     "บวก=ลงทุนแล้วสร้างมูลค่า"),
    ("IRR",
     f"=IRR(B{ncf_y0_row}:B{ncf_rows['Y5']})",
     FMT_PCT,
     "ต้อง > WACC (12%)"),
    ("MIRR (reinvest 5%, finance 12%)",
     f"=MIRR(B{ncf_y0_row}:B{ncf_rows['Y5']},{A['wacc']},0.05)",
     FMT_PCT,
     "IRR ปรับแล้วด้วยอัตรา reinvest จริง"),
    ("Payback Period (base case)",
     "~2.78 ปี",
     None,
     "simple payback (ดูตาราง cumulative CF ด้านบน)"),
    ("รายได้ปีที่ 1",
     f"='P&L'!B{PL['total_revenue']}", FMT_THB, ""),
    ("รายได้ปีที่ 5",
     f"='P&L'!F{PL['total_revenue']}", FMT_THB, ""),
    ("กำไรสุทธิปีที่ 1",
     f"='P&L'!B{PL['net_profit']}", FMT_THB, ""),
    ("กำไรสุทธิปีที่ 5",
     f"='P&L'!F{PL['net_profit']}", FMT_THB, ""),
    ("เงินสดสะสมสิ้นปีที่ 5",
     f"=CashFlow!G{CF['ending_cash']}", FMT_THB, ""),
    ("เงินทุนเริ่มต้นรวม",
     f"={A['founder_cap']}+{A['loan_amount']}", FMT_THB, "equity 2M + debt 1M"),
    ("จำนวน tenant ณ สิ้นปี 5",
     f"='P&L'!F{end_row}", FMT_INT, ""),
]

for label, formula, fmt, note in kpi_items:
    summary.cell(row=row, column=1, value=label).font = TH_FONT
    summary.cell(row=row, column=1).alignment = Alignment(horizontal="left")
    summary.cell(row=row, column=1).border = BORDER
    style_cell(summary.cell(row=row, column=2, value=formula),
               font=TH_BOLD, fill=TOTAL_FILL, fmt=fmt)
    summary.cell(row=row, column=3, value=note).font = TH_SUBTITLE
    summary.cell(row=row, column=3).alignment = Alignment(horizontal="left", wrap_text=True)
    summary.cell(row=row, column=3).border = BORDER
    row += 1

# Reorder sheets
desired_order = [
    "Assumptions",
    "Loan",
    "Assets",
    "SG&A",
    "P&L",
    "CashFlow",
    "BalanceSheet",
    "BreakEven",
    "Sensitivity",
    "Summary",
]
wb._sheets = [wb[name] for name in desired_order]

# Freeze panes
wb["Assumptions"].freeze_panes = "A5"
wb["Loan"].freeze_panes = "A4"
wb["Assets"].freeze_panes = "A4"
wb["SG&A"].freeze_panes = "B5"
wb["P&L"].freeze_panes = "B5"
wb["CashFlow"].freeze_panes = "B5"
wb["BalanceSheet"].freeze_panes = "B5"
wb["BreakEven"].freeze_panes = "A4"
wb["Sensitivity"].freeze_panes = "A4"
wb["Summary"].freeze_panes = "A4"

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"File size: {OUTPUT.stat().st_size:,} bytes")
print()
print("Sheets (ch6 order, 5-year):")
for name in wb.sheetnames:
    print(f"  - {name}")
print()
print("Key formulas:")
print("  Loan!    PMT/Interest/Principal amortization 5 yr")
print("  P&L!     + Interest จาก Loan, + Dividend IF year>=4")
print("  CashFlow! + Financing (equity + loan receipt + principal payment + dividend)")
print("  BS!      + Long-term debt จาก Loan sheet")
print("  Summary! NPV(12%) + IRR + MIRR + Payback + Y5 KPIs")
